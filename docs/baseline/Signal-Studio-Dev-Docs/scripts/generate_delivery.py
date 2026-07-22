#!/usr/bin/env python3
"""Generate the Signal Studio platform development-document baseline.

The script only writes inside Signal-Studio-Dev-Docs and copies source evidence;
it never modifies files under ../references.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
import struct
import textwrap
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw


DATE = "2026-07-22"
VERSION = "V1.0.0"
BASELINE = "BL1.0"
ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parent
REF = REPO / "references"


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def write_json(path: Path, data) -> None:
    write_text(path, json.dumps(data, ensure_ascii=False, indent=2))


def rel_link(path: Path, target: Path) -> str:
    return Path(target).relative_to(path.parent).as_posix()


def content_type(name: str) -> str:
    """Map the engineering document to the closest writing-guideline type."""
    if name == "Signal Studio 开发文档交付基线":
        return "Landing"
    if any(term in name for term in ["架构", "设计说明书", "设计说明", "规范"]):
        return "Conceptual"
    if any(term in name for term in ["计划", "方案"]):
        return "How-to"
    if any(term in name for term in ["已知问题", "错误码", "状态与异常"]):
        return "Troubleshooting"
    return "Reference"


def metadata(number: str, name: str, stage: str, sources: str, change: str, status: str = "已批准") -> str:
    return f"""| 元数据项 | 内容 |
|---|---|
| 文档编号 | {number} |
| 文档名称 | {name} |
| 项目名称 | Signal Studio / Signal Platform |
| 文档版本 | {VERSION} |
| 基线版本 | {BASELINE} |
| 状态 | {status} |
| 内容类型（meta.contentType） | {content_type(name)} |
| 编制日期 | {DATE} |
| 适用阶段 | {stage} |
| 输入来源 | {sources} |
| 本版变更 | {change} |"""


def footer(refs: list[str], unresolved: list[str] | None = None) -> str:
    refs_text = "\n".join(f"- {item}" for item in refs) or "- 无外部参考。"
    items = unresolved or ["无阻断性未决事项；正文中的建议值和待确认项继续按其原状态追踪，不因文档获批而视为已实施。"]
    unresolved_text = "\n".join(f"- {item}" for item in items)
    return f"""## 参考资料

{refs_text}

## 未决事项

{unresolved_text}

## 变更记录

| 版本 | 日期 | 变更 |
|---|---|---|
| {VERSION} | {DATE} | 建立并自动审核通过平台化开发基线，纳入需求、接口、测试和复用边界。 |"""


def md(path: str, number: str, title: str, body: str, stage: str, sources: str,
       change: str, refs: list[str], unresolved: list[str] | None = None,
       status: str = "已批准") -> None:
    content = f"# {title}\n\n{metadata(number, title, stage, sources, change, status)}\n\n{body.strip()}\n\n{footer(refs, unresolved)}"
    write_text(ROOT / path, content)


PREFIX_LIB = {
    "PRJ": "SignalCore", "DAT": "SignalData", "IDX": "SignalData",
    "VIS": "SignalVisualization", "NAV": "SignalVisualization", "SEL": "SignalVisualization",
    "DSP": "SignalDSP", "INS": "SignalWorkbench", "ALG": "SignalModelRuntime",
    "TSK": "SignalTaskRuntime", "EXP": "SignalCore", "LOG": "SignalCore",
    "PLG": "SignalPluginSDK", "SYS": "SignalWorkbench", "PERF": "SignalCompute",
    "REL": "SignalCore", "NUM": "SignalDSP", "SEC": "SignalCore",
    "USA": "SignalWorkbench", "MNT": "SignalCore",
}

PREFIX_PAGE = {
    "PRJ": "UI-P01-001", "DAT": "UI-W01-001", "IDX": "UI-P02-001",
    "VIS": "UI-P02-001", "NAV": "UI-P02-001", "SEL": "UI-P02-001",
    "DSP": "UI-P03-001", "INS": "UI-P03-001", "ALG": "UI-P03-001",
    "TSK": "UI-P04-001", "EXP": "UI-P05-001", "LOG": "UI-P07-001",
    "PLG": "UI-P06-001", "SYS": "UI-P07-001", "USA": "UI-P07-001",
}

PREFIX_API = {
    "PRJ": "API-CORE-006", "DAT": "API-DATA-003", "IDX": "API-DATA-005",
    "VIS": "API-VIS-003", "NAV": "API-VIS-002", "SEL": "API-VIS-006",
    "DSP": "API-DSP-002", "INS": "API-WB-002", "ALG": "API-MODEL-002",
    "TSK": "API-TASK-002", "EXP": "API-CORE-007", "LOG": "API-CORE-005",
    "PLG": "API-PLG-002", "SYS": "API-WB-004", "PERF": "API-COMPUTE-003",
    "REL": "API-CORE-002", "NUM": "API-DSP-001", "SEC": "API-CORE-004",
    "USA": "API-WB-003", "MNT": "API-CORE-001",
}

PREFIX_MS = {
    "PRJ": "MS-01", "DAT": "MS-01", "IDX": "MS-01", "TSK": "MS-01",
    "DSP": "MS-02", "PERF": "MS-02", "NUM": "MS-02",
    "VIS": "MS-03", "NAV": "MS-03", "SEL": "MS-03", "USA": "MS-03",
    "INS": "MS-04", "EXP": "MS-04", "REL": "MS-04",
    "ALG": "MS-06", "PLG": "MS-06", "SYS": "MS-07", "LOG": "MS-07",
    "SEC": "MS-07", "MNT": "MS-07",
}

APP_TARGETS = {
    "SignalCore": "全部应用、Headless CLI",
    "SignalData": "Signal Studio、Signal Generator、Dataset Builder、Dataset Manager、Model Trainer、Inference Studio",
    "SignalDSP": "Signal Studio、Signal Generator、Dataset Builder、Model Trainer、Inference Studio",
    "SignalCompute": "全部计算型应用",
    "SignalTaskRuntime": "全部应用、自动化测试运行器",
    "SignalVisualization": "Signal Studio、Signal Generator、Model Evaluator、Inference Studio",
    "SignalWorkbench": "全部桌面工具",
    "SignalPluginSDK": "全部宿主、Headless CLI、自动化测试运行器",
    "SignalModelRuntime": "Signal Studio、Model Trainer、Model Evaluator、Inference Studio",
    "SignalDataset": "Dataset Builder、Dataset Manager、Model Trainer、Model Evaluator",
}


def parse_requirements() -> list[dict]:
    source = (REF / "01_软件需求规格说明书_SRS_V2.0.md").read_text(encoding="utf-8")
    reqs: list[dict] = []
    seen: set[str] = set()
    nfr_prefixes = {"PERF", "REL", "NUM", "SEC", "USA", "MNT"}
    for line in source.splitlines():
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if not cells or not re.fullmatch(r"SS-[A-Z]+-\d{3}", cells[0]):
            continue
        legacy = cells[0]
        if legacy in seen:
            continue
        seen.add(legacy)
        prefix, seq = legacy.split("-")[1:]
        priority = cells[1] if len(cells) > 1 and re.fullmatch(r"P[012]", cells[1]) else "P0"
        if len(cells) > 2 and re.fullmatch(r"P[012]", cells[1]):
            desc = cells[2]
        elif len(cells) > 2:
            desc = cells[2]
        else:
            desc = cells[1]
        kind = "NFR" if prefix in nfr_prefixes else "FR"
        canonical = f"{kind}-{prefix}-{seq}"
        lib = PREFIX_LIB[prefix]
        reqs.append({
            "id": canonical, "legacy_id": legacy, "name": desc.split("；")[0][:42],
            "description": desc, "source": "SS-SRS-001 V2.0", "layer": "平台" if prefix in nfr_prefixes else "基础库/应用组合",
            "precondition": "已建立有效项目上下文；适用输入契约已通过校验。",
            "input": "用户命令、公共数据契约或任务请求。",
            "processing": desc, "output": "结构化结果、状态或可审计错误。",
            "exception": "返回稳定错误码；不得静默降级或发布半成品。",
            "priority": priority, "acceptance": desc,
            "page": PREFIX_PAGE.get(prefix, ""), "page_applicable": prefix in PREFIX_PAGE,
            "api": PREFIX_API[prefix], "test": "", "milestone": PREFIX_MS[prefix],
            "library": lib, "reuse_apps": APP_TARGETS[lib], "status": "已批准",
        })

    platform_additions = [
        ("FR-PLT-001", "十个公共模块可独立构建、安装、版本化和测试，应用层仅组合公开能力。", "SignalCore", "P0", "MS-00"),
        ("FR-CORE-101", "SignalCore 提供统一 Result/Error、单位、日志、配置、路径、校验、序列化和能力探测。", "SignalCore", "P0", "MS-01"),
        ("FR-DATA-101", "SignalData 提供实数/复数容器、IQ 视图、零拷贝切片、内存映射、分块读取和格式适配。", "SignalData", "P0", "MS-01"),
        ("FR-DSP-101", "SignalDSP 通过 IFftBackend 等接口封装 oneMKL、cuFFT 和可替换适配器，不自行实现成熟 FFT、滤波与重采样内核。", "SignalDSP", "P0", "MS-02"),
        ("FR-COMPUTE-101", "SignalCompute 统一 CPU、SIMD、多线程、CUDA、内存池、设备探测、自动降级和结果一致性。", "SignalCompute", "P0", "MS-02"),
        ("FR-TASK-101", "SignalTaskRuntime 支持任务优先级、DAG、暂停、恢复、取消、进度、重试、超时、历史和崩溃恢复。", "SignalTaskRuntime", "P0", "MS-01"),
        ("FR-VIS-101", "SignalVisualization 提供时域、PSD、瀑布、时频、星座、眼图、联动视口、图层和导出组件。", "SignalVisualization", "P0", "MS-03"),
        ("FR-WB-101", "SignalWorkbench 提供主框架、Dock、Inspector、Task/Result Center、命令、设置、诊断、主题和布局。", "SignalWorkbench", "P0", "MS-03"),
        ("FR-PLG-101", "SignalPluginSDK 提供与宿主 UI 解耦的算法、格式、导入导出、模型和可视化插件契约及示例。", "SignalPluginSDK", "P1", "MS-06"),
        ("FR-MODEL-101", "SignalModelRuntime 以 ONNX Runtime 为默认推理后端，统一模型注册、前后处理、批处理、设备和结果契约。", "SignalModelRuntime", "P1", "MS-06"),
        ("FR-DSET-101", "SignalDataset 提供清单、索引、标签、分片、版本、划分、统计、查询、缓存及 HDF5/WebDataset 适配。", "SignalDataset", "P1", "MS-06"),
        ("NFR-API-101", "公共 API 不暴露 Qt、Eigen、oneMKL、cuFFT、ONNX Runtime 等第三方具体类型。", "SignalCore", "P0", "MS-00"),
        ("NFR-ABI-101", "SDK 采用语义化版本；插件使用版本化 C ABI 入口和稳定 POD 句柄，破坏性变更提升主版本。", "SignalPluginSDK", "P0", "MS-00"),
        ("NFR-DEP-101", "依赖以 vcpkg manifest 和固定 baseline 锁定；外部工具链使用版本、URL、SHA256 和离线缓存清单。", "SignalCore", "P0", "MS-00"),
        ("NFR-TEST-101", "每个公共模块具有单元、契约、兼容、性能和包消费测试；SDK 示例必须在 CI 编译。", "SignalCore", "P0", "MS-00"),
        ("NFR-REUSE-101", "Beta 前至少以 Headless CLI 和第二个薄壳应用验证 Data、DSP、TaskRuntime 与 Visualization 复用。", "SignalWorkbench", "P0", "MS-08"),
    ]
    for rid, desc, lib, priority, ms in platform_additions:
        reqs.append({
            "id": rid, "legacy_id": "", "name": desc.split("，")[0], "description": desc,
            "source": "平台化架构总任务", "layer": "平台/基础库", "precondition": "平台工程已建立。",
            "input": "模块构建、宿主调用或发布请求。", "processing": desc,
            "output": "可独立消费的库、SDK、包或验证证据。", "exception": "违反依赖或兼容约束时在构建/加载阶段失败。",
            "priority": priority, "acceptance": desc + "；对应契约测试通过。",
            "page": "UI-P07-001" if lib in {"SignalWorkbench", "SignalPluginSDK"} else "",
            "page_applicable": lib in {"SignalWorkbench", "SignalPluginSDK"},
            "api": {
                "SignalCore": "API-CORE-001", "SignalData": "API-DATA-003", "SignalDSP": "API-DSP-001",
                "SignalCompute": "API-COMPUTE-001", "SignalTaskRuntime": "API-TASK-002",
                "SignalVisualization": "API-VIS-001", "SignalWorkbench": "API-WB-001",
                "SignalPluginSDK": "API-PLG-001", "SignalModelRuntime": "API-MODEL-001",
                "SignalDataset": "API-DSET-001",
            }[lib], "test": "", "milestone": ms, "library": lib,
            "reuse_apps": APP_TARGETS[lib], "status": "已批准",
        })
    return reqs


def page_catalog() -> list[dict]:
    return [
        {"id": "UI-P01-001", "code": "P01", "name": "项目首页", "owner": "Signal Studio 应用", "reuse": "工作台容器可复用", "targets": "全部桌面工具", "components": "RecentProjectList, SystemStatusCard", "abstract": "是"},
        {"id": "UI-P02-001", "code": "P02", "name": "宽带浏览", "owner": "Signal Studio 应用", "reuse": "图表和视口可复用", "targets": "Signal Generator, Inference Studio", "components": "TimeNavigator, SpectrumView, SpectrogramView", "abstract": "是"},
        {"id": "UI-P03-001", "code": "P03", "name": "Inspector", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部分析工具", "components": "InspectorHost, ProcessingChain", "abstract": "是"},
        {"id": "UI-P04-001", "code": "P04", "name": "任务中心", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部工具", "components": "TaskCenterView", "abstract": "是"},
        {"id": "UI-P05-001", "code": "P05", "name": "结果中心", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部工具", "components": "ResultCenterView", "abstract": "是"},
        {"id": "UI-P06-001", "code": "P06", "name": "插件与模型", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部插件宿主", "components": "PluginManagerView, ModelRegistryView", "abstract": "是"},
        {"id": "UI-P07-001", "code": "P07", "name": "设置与诊断", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部桌面工具", "components": "SettingsView, DiagnosticsView", "abstract": "是"},
        {"id": "UI-W01-001", "code": "W01", "name": "导入信号", "owner": "SignalData + 应用编排", "reuse": "导入框架可复用", "targets": "Studio, Dataset Manager", "components": "ImportWizardHost", "abstract": "是"},
        {"id": "UI-W02-001", "code": "W02", "name": "创建分析通道", "owner": "Signal Studio 应用", "reuse": "Selection 输入组件可复用", "targets": "Inference Studio", "components": "ChannelWizard", "abstract": "部分"},
        {"id": "UI-W03-001", "code": "W03", "name": "导出", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部工具", "components": "ExportDialogHost", "abstract": "是"},
        {"id": "UI-W04-001", "code": "W04", "name": "恢复与资源重定位", "owner": "SignalWorkbench", "reuse": "公共工作台", "targets": "全部项目型工具", "components": "RecoveryDialog", "abstract": "是"},
        {"id": "UI-W05-001", "code": "W05", "name": "加载进度与部分读取", "owner": "SignalTaskRuntime + SignalWorkbench", "reuse": "公共任务界面", "targets": "全部数据工具", "components": "TaskProgressDialog", "abstract": "是"},
    ]


def api_catalog() -> list[dict]:
    rows = [
        ("API-CORE-001", "SignalCore", "signal::core", "VersionInfo version()", "返回平台、API、ABI 与构建版本"),
        ("API-CORE-002", "SignalCore", "signal::core", "Result<T> / Error", "统一成功/失败模型，错误不可丢失"),
        ("API-CORE-003", "SignalCore", "signal::core", "Quantity<Unit, Rep>", "时间、频率、采样率的强类型量"),
        ("API-CORE-004", "SignalCore", "signal::core", "Hash256 hash_file(PathView)", "文件完整性与缓存键"),
        ("API-CORE-005", "SignalCore", "signal::core", "ILogger::write(LogEvent)", "结构化日志，不暴露 spdlog 类型"),
        ("API-CORE-006", "SignalCore", "signal::core", "IWorkspaceStore::load/save", "公共 Workspace 原子持久化"),
        ("API-CORE-007", "SignalCore", "signal::core", "IArtifactStore::commit", "不可变结果与导出制品提交"),
        ("API-DATA-001", "SignalData", "signal::data", "SignalDescriptor::validate()", "验证实/复、dtype、字节序、单位与范围"),
        ("API-DATA-002", "SignalData", "signal::data", "SampleBlockView<T>", "只读零拷贝分块视图"),
        ("API-DATA-003", "SignalData", "signal::data", "IDataSource::read(ReadRequest)", "异步、可取消、帧对齐读取"),
        ("API-DATA-004", "SignalData", "signal::data", "IFormatAdapter::probe/open", "RAW/WAV/HDF5/插件格式适配"),
        ("API-DATA-005", "SignalData", "signal::data", "IMultiResolutionStore::get/put", "概览与瓦片缓存"),
        ("API-COMPUTE-001", "SignalCompute", "signal::compute", "IComputeBackend::capabilities()", "CPU/SIMD/CUDA 能力探测"),
        ("API-COMPUTE-002", "SignalCompute", "signal::compute", "IBufferPool::acquire(BufferSpec)", "受预算约束的主机/设备内存"),
        ("API-COMPUTE-003", "SignalCompute", "signal::compute", "IBackendSelector::select(Workload)", "Auto 选择和可审计降级"),
        ("API-DSP-001", "SignalDSP", "signal::dsp", "IFftBackend::create_plan(FftSpec)", "oneMKL/cuFFT/许可适配器计划"),
        ("API-DSP-002", "SignalDSP", "signal::dsp", "IPsdEstimator::process(SampleBlock)", "窗、ENBW、单位和平均策略明确"),
        ("API-DSP-003", "SignalDSP", "signal::dsp", "IStftProcessor::process(StftRequest)", "分块 STFT 与瓦片输出"),
        ("API-DSP-004", "SignalDSP", "signal::dsp", "IResampler::process(Ratio, SampleBlock)", "抗混叠的成熟库适配"),
        ("API-DSP-005", "SignalDSP", "signal::dsp", "IFilter::process(SampleBlock, State)", "保持跨块状态"),
        ("API-TASK-001", "SignalTaskRuntime", "signal::task", "TaskSpec / ResourceProfile", "冻结的任务、资源和依赖描述"),
        ("API-TASK-002", "SignalTaskRuntime", "signal::task", "ITaskService::submit(TaskSpec)", "提交并返回 TaskHandle"),
        ("API-TASK-003", "SignalTaskRuntime", "signal::task", "TaskHandle::pause/resume/cancel", "幂等控制"),
        ("API-TASK-004", "SignalTaskRuntime", "signal::task", "ITaskObserver::on_event(TaskEvent)", "进度、日志、指标的单一状态源"),
        ("API-VIS-001", "SignalVisualization", "signal::visualization", "IChartView::bind(IDataSeries)", "图表与数据提供者解耦"),
        ("API-VIS-002", "SignalVisualization", "signal::visualization", "ViewportController::set_time/frequency", "多图共享视口"),
        ("API-VIS-003", "SignalVisualization", "signal::visualization", "SpectrumView / SpectrogramView", "统一频率轴与原子提交"),
        ("API-VIS-004", "SignalVisualization", "signal::visualization", "TimeWaveformView / TimeNavigator", "当前视窗与实际读入范围"),
        ("API-VIS-005", "SignalVisualization", "signal::visualization", "ConstellationView / EyeDiagramView", "可复用分析视图"),
        ("API-VIS-006", "SignalVisualization", "signal::visualization", "OverlayModel::selection/measurement", "Selection 与测量覆盖层"),
        ("API-WB-001", "SignalWorkbench", "signal::workbench", "IServiceRegistry::resolve(ServiceId)", "宿主服务发现"),
        ("API-WB-002", "SignalWorkbench", "signal::workbench", "IPanelFactory::create(PanelContext)", "Dock/Inspector/Center 面板"),
        ("API-WB-003", "SignalWorkbench", "signal::workbench", "ICommandRegistry::register_command", "菜单、快捷键与自动化命令"),
        ("API-WB-004", "SignalWorkbench", "signal::workbench", "IDiagnosticsProvider::snapshot", "真实环境和后端诊断"),
        ("API-PLG-001", "SignalPluginSDK", "signal::plugin", "signal_plugin_query_v1(host, out)", "版本化 C ABI 唯一入口"),
        ("API-PLG-002", "SignalPluginSDK", "signal::plugin", "IAlgorithmPlugin::describe/run", "无 UI 依赖的算法插件"),
        ("API-PLG-003", "SignalPluginSDK", "signal::plugin", "IFormatPlugin / IExportPlugin", "格式和导出扩展"),
        ("API-MODEL-001", "SignalModelRuntime", "signal::model", "IModelRegistry::install/resolve", "模型包、版本、哈希和兼容"),
        ("API-MODEL-002", "SignalModelRuntime", "signal::model", "IInferenceSession::run(InferenceRequest)", "ONNX 默认后端、设备可选"),
        ("API-DSET-001", "SignalDataset", "signal::dataset", "IDataset::query(SampleQuery)", "版本化样本索引"),
        ("API-DSET-002", "SignalDataset", "signal::dataset", "IDatasetWriter::append/commit", "分片原子提交"),
        ("API-PY-001", "Python SDK", "signal_platform", "open_signal(path, descriptor)", "NumPy 零拷贝视图"),
        ("API-PY-002", "Python SDK", "signal_platform", "submit_task(spec) / TaskHandle", "任务提交和状态"),
        ("API-PY-003", "Python SDK", "signal_platform", "load_dataset(uri) / infer(model, batch)", "数据集与推理复用"),
    ]
    return [{"id": a, "module": b, "namespace": c, "signature": d, "contract": e,
             "stability": "Stable-1.0" if not a.startswith("API-PY") else "Preview-1.0"} for a, b, c, d, e in rows]


def architecture() -> dict:
    nodes = [
        {"name": "SignalCore", "layer": "基础能力库", "deps": [], "public_namespace": "signal::core"},
        {"name": "SignalCompute", "layer": "计算后端", "deps": ["SignalCore"], "public_namespace": "signal::compute"},
        {"name": "SignalData", "layer": "基础能力库", "deps": ["SignalCore"], "public_namespace": "signal::data"},
        {"name": "SignalTaskRuntime", "layer": "领域服务", "deps": ["SignalCompute", "SignalCore"], "public_namespace": "signal::task"},
        {"name": "SignalDSP", "layer": "基础能力库", "deps": ["SignalData", "SignalCompute", "SignalCore"], "public_namespace": "signal::dsp"},
        {"name": "SignalVisualization", "layer": "公共 UI", "deps": ["SignalData", "SignalTaskRuntime", "SignalCore"], "public_namespace": "signal::visualization"},
        {"name": "SignalModelRuntime", "layer": "领域服务", "deps": ["SignalData", "SignalCompute", "SignalTaskRuntime", "SignalCore"], "public_namespace": "signal::model"},
        {"name": "SignalDataset", "layer": "领域服务", "deps": ["SignalData", "SignalTaskRuntime", "SignalCore"], "public_namespace": "signal::dataset"},
        {"name": "SignalPluginSDK", "layer": "SDK", "deps": ["SignalData", "SignalTaskRuntime", "SignalCore"], "public_namespace": "signal::plugin"},
        {"name": "SignalWorkbench", "layer": "公共 UI", "deps": ["SignalVisualization", "SignalTaskRuntime", "SignalCore"], "public_namespace": "signal::workbench"},
    ]
    apps = ["Signal Studio", "Signal Generator", "Dataset Builder", "Dataset Manager", "Model Trainer", "Model Evaluator", "Inference Studio", "Headless CLI"]
    return {"schema": "signal.platform.architecture/1.0", "nodes": nodes, "apps": apps,
            "rules": ["公共模块不得依赖应用", "禁止循环依赖", "应用不得直接调用具体 FFT/GPU/模型运行库", "公共 API 不暴露第三方类型"]}


def build_tests(reqs: list[dict]) -> list[dict]:
    counters: dict[str, int] = {}
    tests = []
    for req in reqs:
        prefix = req["id"].split("-")[1]
        ttype = {"PERF": "PERF", "REL": "REL", "NUM": "NUM", "SEC": "SEC", "USA": "USA",
                 "ABI": "COMPAT", "API": "COMPAT", "DEP": "BUILD", "TEST": "BUILD", "REUSE": "REUSE"}.get(prefix, "FUNC")
        counters[ttype] = counters.get(ttype, 0) + 1
        tid = f"TC-{ttype}-{counters[ttype]:03d}"
        req["test"] = tid
        tests.append({
            "id": tid, "type": ttype, "name": f"验证 {req['id']}", "requirement": req["id"],
            "priority": req["priority"], "precondition": req["precondition"], "steps": "准备确定性输入；执行契约动作；记录状态、输出和追踪字段；注入一个失败分支。",
            "expected": req["acceptance"], "library": req["library"], "app": req["reuse_apps"],
            "automation": "P0-自动化优先" if req["priority"] == "P0" else "自动化/人工复核", "status": "已设计，待实现执行",
        })
    return tests


def library_capabilities() -> list[dict]:
    capabilities = {
        "CORE": ("SignalCore", ["Result与错误", "单位与版本", "日志与配置", "路径与校验", "序列化与能力探测"]),
        "DATA": ("SignalData", ["信号容器", "描述符与范围", "零拷贝切片", "内存映射", "格式适配与校验"]),
        "DSP": ("SignalDSP", ["FFT/PSD/STFT", "滤波", "重采样", "频移与包络", "统计与同步"]),
        "CMP": ("SignalCompute", ["设备探测", "后端选择", "计划缓存", "内存池", "一致性与降级"]),
        "TASK": ("SignalTaskRuntime", ["队列与优先级", "DAG依赖", "暂停恢复取消", "进度日志指标", "恢复与历史"]),
        "VIS": ("SignalVisualization", ["时域与导航", "频谱与PSD", "瀑布与时频", "星座与眼图", "视口覆盖层导出"]),
        "WB": ("SignalWorkbench", ["应用框架", "Dock与Inspector", "任务结果中心", "命令与设置", "诊断主题DPI"]),
        "PLG": ("SignalPluginSDK", ["插件元数据", "版本化ABI", "服务句柄", "权限隔离", "模板与契约测试"]),
        "MODEL": ("SignalModelRuntime", ["模型注册", "前后处理", "推理会话", "设备与批处理", "结果与性能"]),
        "DSET": ("SignalDataset", ["清单索引", "标签元数据", "分片版本", "划分统计查询", "格式适配缓存"]),
    }
    rows = []
    for short, (lib, caps) in capabilities.items():
        for i, cap in enumerate(caps, 1):
            rows.append({"id": f"LIB-{short}-{i:03d}", "library": lib, "capability": cap,
                         "test_boundary": "模块内单元测试 + 公共契约测试 + 至少一个适配器集成测试",
                         "release": "signal-platform 1.x", "targets": APP_TARGETS[lib]})
    return rows


HEADER_FILL = PatternFill("solid", fgColor="123A5A")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEEF8")
THIN = Side(style="thin", color="B7C9D6")


def style_sheet(ws, widths: dict[str, float] | None = None, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E3EAF0"))
    for col, width in (widths or {}).items():
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def add_notes_sheet(wb: Workbook, title: str, notes: list[str]) -> None:
    ws = wb.create_sheet("说明")
    ws.append([title, "内容"])
    for i, note in enumerate(notes, 1):
        ws.append([i, note])
    style_sheet(ws, {"A": 12, "B": 110})


def save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


def requirements_workbook(reqs: list[dict], tests: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "需求总表"
    headers = ["需求ID", "旧编号", "名称", "描述", "层级", "优先级", "来源", "验收标准", "页面", "接口", "测试", "里程碑", "基础库", "目标应用", "状态"]
    ws.append(headers)
    for r in reqs:
        ws.append([r["id"], r["legacy_id"], r["name"], r["description"], r["layer"], r["priority"], r["source"], r["acceptance"], r["page"], r["api"], r["test"], r["milestone"], r["library"], r["reuse_apps"], r["status"]])
    style_sheet(ws, {"A": 18, "B": 16, "C": 30, "D": 70, "E": 18, "F": 10, "G": 22, "H": 70, "I": 16, "J": 18, "K": 18, "L": 12, "M": 22, "N": 58, "O": 14})
    for col in ["F", "O"]:
        dv = DataValidation(type="list", formula1='"P0,P1,P2"' if col == "F" else '"建议基线,待确认,暂定,已批准"')
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{len(reqs)+1}")

    relations = {
        "需求-页面": ("page", "页面ID"), "需求-接口": ("api", "接口ID"), "需求-测试": ("test", "测试ID"),
        "需求-里程碑": ("milestone", "里程碑ID"), "需求-基础库": ("library", "基础库"), "需求-目标应用": ("reuse_apps", "目标应用"),
    }
    for name, (field, label) in relations.items():
        sheet = wb.create_sheet(name)
        sheet.append(["需求ID", label, "优先级", "状态"])
        for r in reqs:
            if r[field]:
                sheet.append([r["id"], r[field], r["priority"], r["status"]])
        style_sheet(sheet, {"A": 20, "B": 62, "C": 10, "D": 15})

    conflicts = wb.create_sheet("缺口与冲突")
    conflicts.append(["编号", "级别", "冲突/缺口", "依据", "处理", "关闭条件"])
    conflict_rows = [
        ["GAP-001", "P0", "W05 文件大小、初始目标和进度分母存在硬编码冲突", "综合评审 4.1", "生产实现统一由 ReadPlan 派生", "D4 真实元数据测试通过"],
        ["GAP-002", "P0", "旧任务/结果可能被误显示为新数据源当前对象", "综合评审 4.2", "所有对象绑定 DataSourceVersionId", "AT-21 通过"],
        ["GAP-003", "P0", "旧 SRS 导航全文件/叠加标记与最新纯导航语义冲突", "综合评审 4.3", "以 LoadedDataRange 和纯导航为准", "接口、原型、测试一致"],
        ["GAP-004", "P1", "Web 原型为确定性模拟，未证明 Qt/DSP/GPU/模型实现", "证据边界", "原型只作交互证据", "生产测试报告形成"],
        ["GAP-005", "P1", "D1-D4 大规模实体数据未随仓库提供", "输入审计", "生成小型黄金数据和大数据生成脚本", "验收环境生成并登记哈希"],
    ]
    for row in conflict_rows:
        conflicts.append(row)
    style_sheet(conflicts, {"A": 14, "B": 10, "C": 55, "D": 24, "E": 55, "F": 36})

    stats = wb.create_sheet("统计")
    stats.append(["指标", "公式值", "说明"])
    total_row = len(reqs) + 1
    stats_rows = [
        ("需求总数", f"=COUNTA('需求总表'!A2:A{total_row})", "含原 SRS 映射与平台新增需求"),
        ("P0 数", f'=COUNTIF(\'需求总表\'!F2:F{total_row},"P0")', "发布门禁需求"),
        ("P1 数", f'=COUNTIF(\'需求总表\'!F2:F{total_row},"P1")', "承诺/紧随版本需求"),
        ("接口覆盖率", f'=COUNTIF(\'需求总表\'!J2:J{total_row},"<>")/COUNTA(\'需求总表\'!A2:A{total_row})', "需求关联接口比例"),
        ("测试覆盖率", f'=COUNTIF(\'需求总表\'!K2:K{total_row},"<>")/COUNTA(\'需求总表\'!A2:A{total_row})', "需求关联测试设计比例"),
        ("基础库归属覆盖率", f'=COUNTIF(\'需求总表\'!M2:M{total_row},"<>")/COUNTA(\'需求总表\'!A2:A{total_row})', "公共能力归属比例"),
        ("目标应用覆盖率", f'=COUNTIF(\'需求总表\'!N2:N{total_row},"<>")/COUNTA(\'需求总表\'!A2:A{total_row})', "复用目标定义比例"),
    ]
    for row in stats_rows:
        stats.append(row)
    style_sheet(stats, {"A": 28, "B": 22, "C": 55})
    for row in stats.iter_rows(min_row=2):
        if "覆盖率" in str(row[0].value):
            row[1].number_format = "0.0%"
    add_notes_sheet(wb, "需求追踪矩阵", ["所有统计使用 Excel 公式，打开时自动重算。", "页面覆盖率仅对有 UI 适用性的需求统计；非 UI 平台需求通过接口和测试追踪。", "“已设计”不等于测试已执行；执行状态必须由测试报告更新。"])
    save_workbook(wb, ROOT / "01_需求/需求追踪矩阵.xlsx")


def pages_workbook(pages: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "页面清单"
    ws.append(["页面ID", "代码", "名称", "页面归属", "可复用性", "复用目标应用", "依赖公共组件", "需抽象成组件", "原型证据"])
    for p in pages:
        ws.append([p["id"], p["code"], p["name"], p["owner"], p["reuse"], p["targets"], p["components"], p["abstract"], f"页面截图/audit/A{int(p['code'][1:]) if p['code'].startswith('P') else 3:02d}_{p['code']}*.png"])
    style_sheet(ws, {"A": 18, "B": 10, "C": 25, "D": 30, "E": 24, "F": 48, "G": 55, "H": 18, "I": 38})
    dv = DataValidation(type="list", formula1='"是,部分,否"')
    ws.add_data_validation(dv)
    dv.add(f"H2:H{len(pages)+1}")
    stats = wb.create_sheet("统计")
    stats.append(["指标", "公式"])
    stats.append(["页面/流程总数", f"=COUNTA('页面清单'!A2:A{len(pages)+1})"])
    stats.append(["需抽象组件数", f'=COUNTIF(\'页面清单\'!H2:H{len(pages)+1},"是")'])
    stats.append(["公共工作台归属数", f'=COUNTIF(\'页面清单\'!D2:D{len(pages)+1},"*SignalWorkbench*")'])
    style_sheet(stats, {"A": 30, "B": 18})
    add_notes_sheet(wb, "页面与流程界面清单", ["P01-P07 为持久主页面，W01-W05 为受控流程界面。", "公共工作台页面不得硬编码 Signal Studio 品牌或数据类型。", "原型截图是交互证据，不是生产完成证据。"])
    save_workbook(wb, ROOT / "02_原型设计/页面清单.xlsx")


def tests_workbook(tests: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "功能测试"
    headers = ["测试ID", "类型", "名称", "需求", "优先级", "前置条件", "步骤", "预期", "基础库", "复用宿主", "自动化", "状态"]
    ws.append(headers)
    for t in tests:
        ws.append([t[k] for k in ["id", "type", "name", "requirement", "priority", "precondition", "steps", "expected", "library", "app", "automation", "status"]])
    style_sheet(ws, {"A": 18, "B": 12, "C": 28, "D": 18, "E": 10, "F": 45, "G": 70, "H": 70, "I": 24, "J": 55, "K": 20, "L": 22})
    for col, formula in [("E", '"P0,P1,P2"'), ("L", '"已设计，待实现执行,执行中,通过,失败,阻塞"')]:
        dv = DataValidation(type="list", formula1=formula)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{len(tests)+1}")
    categories = [
        ("基础库测试", ["SignalCore", "SignalData", "SignalDSP", "SignalCompute", "SignalTaskRuntime", "SignalVisualization", "SignalWorkbench", "SignalModelRuntime", "SignalDataset"]),
        ("SDK测试", ["SignalPluginSDK", "SignalCore"]),
        ("插件兼容测试", ["SignalPluginSDK"]),
        ("多应用复用测试", ["SignalData", "SignalDSP", "SignalTaskRuntime", "SignalVisualization", "SignalWorkbench"]),
    ]
    for name, libs in categories:
        sheet = wb.create_sheet(name)
        sheet.append(headers)
        selected = [t for t in tests if t["library"] in libs][:60]
        for t in selected:
            sheet.append([t[k] for k in ["id", "type", "name", "requirement", "priority", "precondition", "steps", "expected", "library", "app", "automation", "status"]])
        style_sheet(sheet, {"A": 18, "B": 12, "C": 28, "D": 18, "E": 10, "F": 40, "G": 65, "H": 65, "I": 24, "J": 50, "K": 18, "L": 22})
    stats = wb.create_sheet("统计")
    stats.append(["指标", "值"])
    n = len(tests) + 1
    stats.append(["测试总数", f"=COUNTA('功能测试'!A2:A{n})"])
    stats.append(["P0 测试数", f'=COUNTIF(\'功能测试\'!E2:E{n},"P0")'])
    stats.append(["自动化优先数", f'=COUNTIF(\'功能测试\'!K2:K{n},"*自动化*")'])
    style_sheet(stats, {"A": 30, "B": 18})
    add_notes_sheet(wb, "测试用例基线", ["本工作簿是可执行用例设计，不表示已通过。", "P0/P1 需求均关联至少一个测试 ID。", "实际执行必须记录环境、输入哈希、结果和缺陷号。"])
    save_workbook(wb, ROOT / "06_测试与验收/功能测试用例.xlsx")


def wbs_and_risk_workbooks(libs: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"
    headers = ["WBS", "工作包", "所属基础库", "所属应用", "可复用目标", "API/ABI影响", "SDK影响", "第三方依赖", "独立测试", "发布包", "里程碑", "前置WBS", "估算人周", "验收输出", "状态"]
    ws.append(headers)
    rows = []
    for i, cap in enumerate(libs, 1):
        ms = "MS-01" if cap["library"] in {"SignalCore", "SignalData", "SignalTaskRuntime"} else "MS-02" if cap["library"] in {"SignalDSP", "SignalCompute"} else "MS-03" if cap["library"] in {"SignalVisualization", "SignalWorkbench"} else "MS-06"
        rows.append([f"WBS-1.{i:02d}", cap["capability"], cap["library"], "平台", cap["targets"], "公开契约评审", "可能", "见依赖锁定", cap["test_boundary"], f"{cap['library'].lower()}-dev/runtime", ms, "WBS-0.01" if i > 1 else "", 1.5, f"{cap['id']} API、测试和包", "计划中"])
    app_rows = [
        ["WBS-2.01", "Signal Studio 项目与导入薄壳", "SignalData/Workbench", "Signal Studio", "复用导入和项目基础设施", "无公共ABI变更", "否", "Qt 6", "应用集成测试", "signal-studio", "MS-04", "WBS-1.01", 4, "P01/W01/W05", "计划中"],
        ["WBS-2.02", "宽带、窄带与联动分析编排", "DSP/Visualization", "Signal Studio", "Inference Studio 可复用", "API消费", "否", "oneMKL/CUDA", "AT-05/14/23", "signal-studio", "MS-05", "WBS-2.01", 8, "P02/P03", "计划中"],
        ["WBS-3.01", "第二宿主复用验证", "Data/DSP/Task/Visualization", "Signal Generator 薄壳", "证明非单应用平台", "兼容门禁", "是", "同平台锁", "多应用复用测试", "signal-generator-smoke", "MS-08", "WBS-1.50", 3, "复用报告与零复制代码审计", "计划中"],
    ]
    for row in rows + app_rows:
        ws.append(row)
    style_sheet(ws, {"A": 14, "B": 34, "C": 28, "D": 20, "E": 48, "F": 20, "G": 12, "H": 28, "I": 42, "J": 30, "K": 12, "L": 14, "M": 12, "N": 42, "O": 14})
    dv = DataValidation(type="list", formula1='"计划中,进行中,已完成,阻塞"')
    ws.add_data_validation(dv)
    dv.add(f"O2:O{ws.max_row}")
    total = ws.max_row
    stats = wb.create_sheet("统计")
    stats.append(["指标", "公式"])
    stats.append(["工作包总数", f"=COUNTA(WBS!A2:A{total})"])
    stats.append(["总估算人周", f"=SUM(WBS!M2:M{total})"])
    stats.append(["基础能力工作包", f'=COUNTIF(WBS!D2:D{total},"平台")'])
    style_sheet(stats, {"A": 30, "B": 18})
    add_notes_sheet(wb, "工作分解结构", ["基础能力工作包优先于应用功能。", "估算为建议基线，不代表已批准资源承诺。", "工作包完成需同时具备 API、测试、包和文档。"])
    save_workbook(wb, ROOT / "07_项目计划/工作分解结构WBS.xlsx")

    risks = [
        ("RISK-001", "高", "FFTW GPL 与闭源交付冲突", "法律/依赖", "选 oneMKL；FFTW 仅 GPL 配置", "许可证审计门禁", "架构负责人", "MS-00"),
        ("RISK-002", "高", "跨数据源对象污染导致错误结论", "数据正确性", "DataSourceVersionId 强制外键", "AT-21", "数据负责人", "MS-04"),
        ("RISK-003", "高", "部分读取被误认为完整文件", "数据正确性", "任务与数据状态正交", "AT-19", "任务负责人", "MS-01"),
        ("RISK-004", "高", "CPU/GPU 数值或归一化不一致", "算法", "黄金数据、误差预算、后端一致性", "NUM/性能测试", "DSP负责人", "MS-02"),
        ("RISK-005", "高", "SDK ABI 被编译器或第三方类型破坏", "兼容", "C ABI、PIMPL、abi-compliance-checker", "兼容矩阵", "SDK负责人", "MS-06"),
        ("RISK-006", "高", "平台被 Signal Studio 特例反向污染", "架构", "依赖规则和第二宿主门禁", "架构静态校验", "架构负责人", "MS-08"),
        ("RISK-007", "中", "超大文件索引耗尽磁盘", "性能", "配额/LRU/可恢复瓦片", "故障注入", "数据负责人", "MS-04"),
        ("RISK-008", "中", "CUDA/驱动/ONNX EP 组合不兼容", "部署", "能力探测和CPU回退", "设备矩阵", "计算负责人", "MS-07"),
        ("RISK-009", "中", "Qt Graphs/QCustomPlot 许可证或性能不满足", "UI依赖", "适配层；基准后冻结", "绘制基准", "UI负责人", "MS-03"),
        ("RISK-010", "中", "大规模 D1-D4 数据缺失", "测试", "脚本生成、哈希登记、受控存储", "数据清单", "测试负责人", "MS-00"),
        ("RISK-011", "中", "插件崩溃或越权", "安全", "签名/权限/进程隔离/安全模式", "故障注入", "SDK负责人", "MS-06"),
        ("RISK-012", "中", "高 DPI 紧凑布局命中区不足", "可用性", "28px透明命中和DPI矩阵", "AT-24", "UI负责人", "MS-03"),
    ]
    wb2 = Workbook()
    rs = wb2.active
    rs.title = "风险清单"
    rs.append(["风险ID", "等级", "风险", "类别", "应对", "验证/触发", "责任角色", "关闭里程碑", "状态"])
    for row in risks:
        rs.append([*row, "开放"])
    style_sheet(rs, {"A": 14, "B": 10, "C": 42, "D": 18, "E": 50, "F": 35, "G": 18, "H": 14, "I": 12})
    dv = DataValidation(type="list", formula1='"开放,缓解中,已接受,已关闭"')
    rs.add_data_validation(dv)
    dv.add(f"I2:I{len(risks)+1}")
    summary = wb2.create_sheet("统计")
    summary.append(["指标", "公式"])
    summary.append(["风险总数", f"=COUNTA('风险清单'!A2:A{len(risks)+1})"])
    summary.append(["高风险数", f'=COUNTIF(\'风险清单\'!B2:B{len(risks)+1},"高")'])
    summary.append(["开放风险数", f'=COUNTIF(\'风险清单\'!I2:I{len(risks)+1},"开放")'])
    style_sheet(summary, {"A": 28, "B": 18})
    add_notes_sheet(wb2, "风险登记册", ["高风险每个里程碑复评。", "关闭必须附验证证据，不以口头确认代替。", "许可证、ABI、来源错配均视作发布门禁。"])
    save_workbook(wb2, ROOT / "07_项目计划/风险清单.xlsx")


def svg_document(background: str | None, foreground: str, accent: str, horizontal: bool = False,
                 monochrome: bool = False, app_icon: bool = False) -> str:
    width = 720 if horizontal else 256
    height = 256
    bg = f'<rect width="{width}" height="{height}" rx="{48 if app_icon else 0}" fill="{background}"/>' if background else ""
    fg = foreground
    ac = foreground if monochrome else accent
    mark = f"""
  <g transform="translate({36 if not horizontal else 26} 32)">
    <path d="M8 32 H176 V86 C137 56 102 51 72 65 C48 76 31 96 8 113 Z" fill="{fg}"/>
    <path d="M10 76 C40 76 49 115 75 115 C102 115 112 55 139 55 C166 55 168 92 190 92" fill="none" stroke="{ac}" stroke-width="13" stroke-linecap="round"/>
    <path d="M14 151 C41 130 66 133 88 150 C111 169 133 175 166 148" fill="none" stroke="{fg}" stroke-width="20" stroke-linecap="round"/>
    <g fill="{ac}">
      <rect x="12" y="188" width="12" height="20"/><rect x="34" y="178" width="12" height="30"/>
      <rect x="56" y="164" width="12" height="44"/><rect x="78" y="181" width="12" height="27"/>
      <rect x="100" y="190" width="12" height="18"/><rect x="122" y="176" width="12" height="32"/>
      <rect x="144" y="159" width="12" height="49"/><rect x="166" y="174" width="12" height="34"/>
    </g>
  </g>"""
    word = ""
    if horizontal:
        word = f"""<g transform="translate(250 76)"><text x="0" y="46" fill="{fg}" font-family="Arial, sans-serif" font-size="48" font-weight="700" letter-spacing="1">Signal Studio</text><text x="2" y="82" fill="{ac}" font-family="Arial, sans-serif" font-size="19" letter-spacing="4">SIGNAL ANALYSIS WORKBENCH</text></g>"""
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-labelledby="title desc">
  <title id="title">Signal Studio 品牌标识</title>
  <desc id="desc">由频谱、波形和 S 形负空间组成的原创几何标识</desc>
  <metadata>SPDX-License-Identifier: LicenseRef-Signal-Studio-Project; Created 2026-07-22; no third-party artwork.</metadata>
  {bg}{mark}{word}
</svg>"""


def icon_svg(name: str, path_data: str) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" role="img" aria-labelledby="title">
  <title id="title">{name}</title>
  <metadata>SPDX-License-Identifier: LicenseRef-Signal-Studio-Project; original line icon.</metadata>
  {path_data}
</svg>"""


def generate_assets(concept: Path | None) -> None:
    logo_dir = ROOT / "03_UI规范/Logo"
    icon_dir = ROOT / "03_UI规范/SVG图标"
    image_dir = ROOT / "03_UI规范/Images"
    variants = {
        "SignalStudio_Logo_Primary.svg": svg_document(None, "#0B1324", "#20D3EE"),
        "SignalStudio_Logo_Dark.svg": svg_document("#0B1324", "#F4FAFF", "#20D3EE"),
        "SignalStudio_Logo_Light.svg": svg_document("#F7FBFF", "#0B1324", "#2563EB"),
        "SignalStudio_Logo_Monochrome.svg": svg_document(None, "#111827", "#111827", monochrome=True),
        "SignalStudio_Logo_Horizontal.svg": svg_document(None, "#0B1324", "#20D3EE", horizontal=True),
        "SignalStudio_Mark.svg": svg_document(None, "#0B1324", "#20D3EE"),
        "SignalStudio_AppIcon.svg": svg_document("#0B1324", "#F4FAFF", "#20D3EE", app_icon=True),
    }
    for name, content in variants.items():
        write_text(logo_dir / name, content)
    icons = {
        "import": '<path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M4 20h16"/>',
        "task": '<rect x="4" y="4" width="16" height="16" rx="2"/><path d="M8 9h8M8 13h5M8 17h7"/>',
        "result": '<path d="M4 19V5h16v14z"/><path d="m7 15 3-4 3 2 4-6"/>',
        "model": '<rect x="3" y="3" width="18" height="18" rx="4"/><circle cx="9" cy="9" r="2"/><circle cx="15" cy="15" r="2"/><path d="m10.5 10.5 3 3"/>',
        "settings": '<circle cx="12" cy="12" r="3"/><path d="M12 2v3M12 19v3M2 12h3M19 12h3M4.9 4.9 7 7M17 17l2.1 2.1M19.1 4.9 17 7M7 17l-2.1 2.1"/>',
        "waveform": '<path d="M2 12h3l2-7 4 14 3-10 3 6h5"/>',
        "spectrum": '<path d="M3 20V4M3 20h18"/><path d="M5 17c3 0 3-10 6-10s3 8 5 8 2-4 5-4"/>',
        "waterfall": '<path d="M3 5h18M3 9h14M3 13h18M3 17h10M3 21h18"/>',
        "zoom": '<circle cx="10" cy="10" r="6"/><path d="m15 15 6 6M10 7v6M7 10h6"/>',
        "select": '<path d="M4 3v18M4 3h18M4 21h18"/><path d="m10 9 8 3-4 2-2 4z"/>',
        "warning": '<path d="M12 3 2 21h20z"/><path d="M12 9v5M12 18h.01"/>',
        "gpu": '<rect x="3" y="6" width="18" height="12" rx="2"/><rect x="8" y="9" width="8" height="6"/><path d="M7 3v3M12 3v3M17 3v3M7 18v3M12 18v3M17 18v3"/>',
    }
    for name, path in icons.items():
        write_text(icon_dir / f"ss-{name}.svg", icon_svg(name, path))

    # Rasterize the vector language deterministically using Pillow.
    for size in [16, 24, 32, 48, 64, 128, 256, 512]:
        img = Image.new("RGBA", (size, size), (11, 19, 36, 255))
        d = ImageDraw.Draw(img)
        s = size / 256
        d.rounded_rectangle((1, 1, size - 2, size - 2), radius=max(2, int(42*s)), fill=(11, 19, 36, 255))
        points = [(30*s, 78*s), (72*s, 78*s), (92*s, 126*s), (130*s, 52*s), (170*s, 110*s), (226*s, 80*s)]
        d.line(points, fill=(32, 211, 238, 255), width=max(2, int(13*s)), joint="curve")
        d.arc((38*s, 80*s, 218*s, 200*s), 20, 158, fill=(244, 250, 255, 255), width=max(2, int(20*s)))
        for i, h in enumerate([28, 44, 64, 38, 24, 42, 70, 50]):
            x = (28 + i*25)*s
            d.rectangle((x, (220-h)*s, x+10*s, 220*s), fill=(59, 130, 246, 255))
        img.save(logo_dir / f"SignalStudio_AppIcon_{size}.png")
    Image.open(logo_dir / "SignalStudio_AppIcon_256.png").save(logo_dir / "SignalStudio_AppIcon.ico", sizes=[(16,16),(32,32),(48,48),(64,64),(128,128),(256,256)])
    if concept and concept.exists():
        image_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(concept, image_dir / "SignalStudio_LogoConcept_AI.png")
    asset_manifest = {
        "schema": "signal.assets/1.0", "created": DATE,
        "license": "LicenseRef-Signal-Studio-Project",
        "originality": "All SVG/PNG production assets are generated for this delivery and contain no third-party artwork.",
        "concept": "AI raster concept is retained as design evidence; production assets are deterministic SVG/PNG.",
        "files": sorted([p.relative_to(ROOT / "03_UI规范").as_posix() for p in (ROOT / "03_UI规范").rglob("*.svg")]) + sorted([p.relative_to(ROOT / "03_UI规范").as_posix() for p in logo_dir.glob("*.png")]),
    }
    write_json(ROOT / "03_UI规范/asset-manifest.json", asset_manifest)


def generate_test_data_script_and_data() -> None:
    script = '''#!/usr/bin/env python3
"""Generate deterministic smoke/golden signal files without implementing FFT.

Expected peaks are analytical. Numeric FFT/PSD validation must use the selected
oneMKL/cuFFT adapter or an approved external reference implementation.
"""
import argparse, hashlib, json, math, random, struct
from pathlib import Path

def sha256(path):
    h=hashlib.sha256()
    with path.open("rb") as f:
        for b in iter(lambda:f.read(1<<20), b""): h.update(b)
    return h.hexdigest()

def real_f32(path, frames, fs, tone):
    with path.open("wb") as f:
        for n in range(frames):
            f.write(struct.pack("<f", 0.7*math.sin(2*math.pi*tone*n/fs)))

def complex_sc16(path, frames, fs, tone, noise=0.0, seed=240722):
    rng=random.Random(seed)
    with path.open("wb") as f:
        for n in range(frames):
            p=2*math.pi*tone*n/fs
            i=max(-32768,min(32767,round(24000*math.cos(p)+noise*rng.gauss(0,1))))
            q=max(-32768,min(32767,round(24000*math.sin(p)+noise*rng.gauss(0,1))))
            f.write(struct.pack("<hh", i, q))

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--out", default="generated"); ap.add_argument("--profile", choices=["smoke","D1","D2","D3"], default="smoke"); ap.add_argument("--frames", type=int); a=ap.parse_args()
    out=Path(a.out); out.mkdir(parents=True, exist_ok=True)
    profiles={"smoke":200000,"D1":2500000000,"D2":25000000000,"D3":1250000000}; frames=a.frames or profiles[a.profile]
    if a.profile!="smoke" and a.frames is None:
        raise SystemExit("Large profiles require explicit --frames after capacity review; values are documented targets, not automatic allocations.")
    items=[]
    p=out/"D0_real_f32_10k.bin"; real_f32(p,min(frames,10000),100000,12500); items.append((p,"real",100000,12500,"float32"))
    p=out/"D0_complex_sc16.bin"; complex_sc16(p,frames,1000000,125000,120); items.append((p,"complex",1000000,125000,"int16"))
    bad=out/"D0_corrupt_tail.sc16"; bad.write_bytes((out/"D0_complex_sc16.bin").read_bytes()[:4096]+b"\\x01"); items.append((bad,"complex",1000000,125000,"int16-corrupt-tail"))
    manifest={"schema":"signal.testdata/1.0","license":"CC0-1.0","profile":a.profile,"files":[]}
    for p,kind,fs,tone,dtype in items:
        manifest["files"].append({"path":p.name,"bytes":p.stat().st_size,"sha256":sha256(p),"signalKind":kind,"sampleRateHz":fs,"expectedToneHz":tone,"dtype":dtype})
    (out/"manifest.json").write_text(json.dumps(manifest,indent=2),encoding="utf-8")
if __name__=="__main__": main()
'''
    write_text(ROOT / "06_测试与验收/测试数据/generate_test_data.py", script)
    data_dir = ROOT / "06_测试与验收/测试数据/generated"
    data_dir.mkdir(parents=True, exist_ok=True)
    # Generate the default smoke set in-process with the same contract.
    real_path = data_dir / "D0_real_f32_10k.bin"
    with real_path.open("wb") as f:
        for n in range(10_000):
            f.write(struct.pack("<f", 0.7 * math.sin(2 * math.pi * 12_500 * n / 100_000)))
    complex_path = data_dir / "D0_complex_sc16.bin"
    import random
    rng = random.Random(240722)
    with complex_path.open("wb") as f:
        for n in range(200_000):
            p = 2 * math.pi * 125_000 * n / 1_000_000
            i = max(-32768, min(32767, round(24000 * math.cos(p) + 120*rng.gauss(0, 1))))
            q = max(-32768, min(32767, round(24000 * math.sin(p) + 120*rng.gauss(0, 1))))
            f.write(struct.pack("<hh", i, q))
    corrupt = data_dir / "D0_corrupt_tail.sc16"
    corrupt.write_bytes(complex_path.read_bytes()[:4096] + b"\x01")
    files = []
    for p, kind, fs, tone, dtype in [
        (real_path, "real", 100_000, 12_500, "float32"),
        (complex_path, "complex", 1_000_000, 125_000, "int16"),
        (corrupt, "complex", 1_000_000, 125_000, "int16-corrupt-tail"),
    ]:
        files.append({"path": p.name, "bytes": p.stat().st_size,
                      "sha256": hashlib.sha256(p.read_bytes()).hexdigest(),
                      "signalKind": kind, "sampleRateHz": fs, "expectedToneHz": tone, "dtype": dtype})
    write_json(data_dir / "manifest.json", {"schema": "signal.testdata/1.0", "license": "CC0-1.0", "profile": "smoke", "files": files})
    write_text(ROOT / "06_测试与验收/测试数据/LICENSE.txt", "SPDX-License-Identifier: CC0-1.0\n本目录中的合成测试数据由生成脚本确定性产生，可用于测试和再分发。")


def copy_evidence() -> None:
    source_dir = ROOT / "02_原型设计/原型源文件"
    source_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(REF / "Signal Studio交互原型.html", source_dir / "Signal Studio交互原型_基线归档.html")
    audit_dir = ROOT / "02_原型设计/页面截图/audit"
    audit_dir.mkdir(parents=True, exist_ok=True)
    for p in (REF / "assets/audit").glob("*.png"):
        shutil.copy2(p, audit_dir / p.name)
    hashes = []
    for p in [source_dir / "Signal Studio交互原型_基线归档.html", *sorted(audit_dir.glob("*.png"))]:
        hashes.append({"path": p.relative_to(ROOT).as_posix(), "bytes": p.stat().st_size, "sha256": hashlib.sha256(p.read_bytes()).hexdigest(), "source": "references（只读复制）"})
    write_json(source_dir / "source-manifest.json", {"schema": "signal.prototype-evidence/1.0", "copied": DATE, "files": hashes, "statement": "文件为只读证据副本；未修改原始材料。"})


def dependency_files() -> list[dict]:
    baseline = "82b6bc886d7b0f8342e34babc2e0b8943f79b0e1"
    versions = {
        "qtbase": ("6.11.1#1", "LGPL-3.0-only OR GPL-3.0-only OR LicenseRef-Qt-Commercial"),
        "qttools": ("6.11.1", "LGPL-3.0-only OR GPL-3.0-only OR LicenseRef-Qt-Commercial"),
        "qtgraphs": ("6.11.1", "GPL-3.0-only OR LicenseRef-Qt-Commercial"),
        "intel-mkl": ("2025.2.0#1", "LicenseRef-Intel-Simplified-Software-License"),
        "eigen3": ("5.0.1", "MPL-2.0"), "tbb": ("2023.1.0", "Apache-2.0"),
        "hdf5": ("2.1.1#1", "BSD-3-Clause"), "nlohmann-json": ("3.12.0#2", "MIT"),
        "tomlplusplus": ("3.4.0#1", "MIT"), "pybind11": ("3.0.4", "BSD-3-Clause"),
        "onnxruntime": ("1.23.2#1", "MIT"), "spdlog": ("1.17.0#1", "MIT"),
        "fmt": ("12.2.0#1", "MIT"), "gtest": ("1.17.0#3", "BSD-3-Clause"),
        "benchmark": ("1.9.5", "Apache-2.0"), "liquid-dsp": ("1.8.0", "MIT"),
        "fftw3": ("3.3.11", "GPL-2.0-or-later"), "openblas": ("0.3.33", "BSD-3-Clause"),
        "libtorch": ("2.12.0", "BSD-3-Clause"),
    }
    selected = {"qtbase", "qttools", "intel-mkl", "eigen3", "tbb", "hdf5", "nlohmann-json", "tomlplusplus", "pybind11", "onnxruntime", "spdlog", "fmt", "gtest", "benchmark"}
    official = {
        "qtbase": "https://code.qt.io/cgit/qt/qtbase.git/", "qttools": "https://code.qt.io/cgit/qt/qttools.git/", "qtgraphs": "https://code.qt.io/cgit/qt/qtgraphs.git/",
        "intel-mkl": "https://www.intel.com/content/www/us/en/developer/tools/oneapi/onemkl.html", "eigen3": "https://gitlab.com/libeigen/eigen", "tbb": "https://github.com/uxlfoundation/oneTBB",
        "hdf5": "https://github.com/HDFGroup/hdf5", "nlohmann-json": "https://github.com/nlohmann/json", "tomlplusplus": "https://github.com/marzer/tomlplusplus",
        "pybind11": "https://github.com/pybind/pybind11", "onnxruntime": "https://github.com/microsoft/onnxruntime", "spdlog": "https://github.com/gabime/spdlog",
        "fmt": "https://github.com/fmtlib/fmt", "gtest": "https://github.com/google/googletest", "benchmark": "https://github.com/google/benchmark",
        "liquid-dsp": "https://github.com/jgaeddert/liquid-dsp", "fftw3": "https://www.fftw.org/", "openblas": "https://github.com/OpenMathLib/OpenBLAS", "libtorch": "https://pytorch.org/",
    }
    deps = []
    for name, (version, spdx) in versions.items():
        deps.append({"name": name, "version": version, "spdx": spdx, "selected": name in selected,
                     "official_url": official[name], "lock": f"vcpkg baseline {baseline}",
                     "verification": "baseline.json sha256:9f3b13f9a142969a043a921f544c637c54be46b06eabb9025b7e5c28b908af58",
                     "decision": "选用" if name in selected else "候选/延后",
                     "notes": "FFTW 仅 GPL 配置，不进入默认闭源分发" if name == "fftw3" else "通过适配器隔离"})
    deps.append({"name": "CUDA Toolkit/cuFFT", "version": "12.8.1", "spdx": "LicenseRef-NVIDIA-CUDA-EULA", "selected": False,
                 "official_url": "https://developer.download.nvidia.com/compute/cuda/12.8.1/network_installers/cuda_12.8.1_windows_network.exe",
                 "lock": "external toolchain", "verification": "sha256:779bee8ff557255c1cf5f36e0230f081675b9bb41e44be38839920cd5209bdeb",
                 "decision": "可选 GPU 后端", "notes": "需接受 NVIDIA EULA；不自动安装"})
    manifest = {
        "name": "signal-platform-dependencies", "version-string": "1.0.0",
        "builtin-baseline": baseline,
        "dependencies": ["qtbase", "qttools", "intel-mkl", "eigen3", "tbb", "hdf5", "nlohmann-json", "tomlplusplus", "pybind11", "onnxruntime", "spdlog", "fmt", "gtest", "benchmark"],
        "features": {"gpu-cuda": {"description": "CUDA/cuFFT is an externally locked toolchain; see dependency-lock.json", "dependencies": []},
                     "telecom": {"description": "Optional liquid-dsp adapter", "dependencies": ["liquid-dsp"]}},
    }
    lock_dir = ROOT / "08_参考资料/依赖锁定"
    write_json(lock_dir / "vcpkg.json", manifest)
    write_json(lock_dir / "dependency-lock.json", {"schema": "signal.dependencies/1.0", "generated": DATE, "vcpkg_baseline": baseline,
                                                     "vcpkg_archive_sha256": "550800632708a561c82412ee69e227c261d0ac8bc381eee09d123014528ae97a",
                                                     "vcpkg_archive_bytes": 5332790, "dependencies": deps})
    write_text(lock_dir / "vcpkg-baseline.sha256", "9f3b13f9a142969a043a921f544c637c54be46b06eabb9025b7e5c28b908af58  baseline.json")
    fetch = '''param([string]$Destination = "$PSScriptRoot\\cache")
$ErrorActionPreference = "Stop"
$commit = "82b6bc886d7b0f8342e34babc2e0b8943f79b0e1"
$url = "https://github.com/microsoft/vcpkg/archive/$commit.tar.gz"
$sha = "550800632708a561c82412ee69e227c261d0ac8bc381eee09d123014528ae97a"
New-Item -ItemType Directory -Force $Destination | Out-Null
$archive = Join-Path $Destination "vcpkg-$commit.tar.gz"
Invoke-WebRequest -UseBasicParsing $url -OutFile $archive
if ((Get-FileHash $archive -Algorithm SHA256).Hash.ToLowerInvariant() -ne $sha) { throw "vcpkg archive checksum mismatch" }
Write-Host "Verified $archive"
'''
    write_text(ROOT / "scripts/fetch_dependencies.ps1", fetch)
    external = '''param([string]$Destination = "$PSScriptRoot\\offline-cache")
$ErrorActionPreference = "Stop"
New-Item -ItemType Directory -Force $Destination | Out-Null
$url = "https://developer.download.nvidia.com/compute/cuda/12.8.1/network_installers/cuda_12.8.1_windows_network.exe"
$out = Join-Path $Destination "cuda_12.8.1_windows_network.exe"
Invoke-WebRequest -UseBasicParsing $url -OutFile $out
$expected = "779bee8ff557255c1cf5f36e0230f081675b9bb41e44be38839920cd5209bdeb"
if ((Get-FileHash $out -Algorithm SHA256).Hash.ToLowerInvariant() -ne $expected) { throw "CUDA installer checksum mismatch" }
Write-Host "Verified. Installation is manual and requires acceptance of the NVIDIA CUDA EULA."
'''
    write_text(ROOT / "scripts/fetch_external_dependencies.ps1", external)
    return deps


def requirement_table(reqs: list[dict], limit: int | None = None) -> str:
    rows = ["| 编号 | 旧编号 | 优先级 | 层级 | 需求摘要 | 验收 | 页面 | 接口 | 测试 | 里程碑 | 复用 |",
            "|---|---|---|---|---|---|---|---|---|---|---|"]
    for r in (reqs[:limit] if limit else reqs):
        rows.append(f"| {r['id']} | {r['legacy_id'] or '新增'} | {r['priority']} | {r['layer']} | {r['description']} | {r['acceptance']} | {r['page'] or '不适用'} | {r['api']} | {r['test']} | {r['milestone']} | {r['reuse_apps']} |")
    return "\n".join(rows)


def api_table(apis: list[dict]) -> str:
    rows = ["| API ID | 模块 | 命名空间 | 签名/类型 | 契约 | 稳定性 |", "|---|---|---|---|---|---|"]
    for a in apis:
        rows.append(f"| {a['id']} | {a['module']} | `{a['namespace']}` | `{a['signature']}` | {a['contract']} | {a['stability']} |")
    return "\n".join(rows)


def generate_docs(reqs: list[dict], pages: list[dict], apis: list[dict], tests: list[dict], libs: list[dict], deps: list[dict]) -> None:
    common_refs = ["原始材料：`../references/`（交付目录之外，只读输入）", "平台任务提示词（平台化架构版）"]
    # Root and delivery control.
    md("README.md", "SS-DOC-INDEX-001", "Signal Studio 开发文档交付基线",
       """## 1. 交付结论

本目录是经自动审核通过的正式文档基线。架构以 **Signal Platform 可复用基础能力平台** 为主产品，Signal Studio 是第一个应用薄壳；任何公共模块不得反向依赖应用。

## 2. 推荐阅读顺序

1. [自动审核批准记录](00_交付说明/自动审核批准记录.md)、[执行报告](执行报告.md)与[交付物清单](00_交付说明/交付物清单.md)；
2. [SRS](01_需求/软件需求规格说明书.md)与[需求追踪矩阵](01_需求/需求追踪矩阵.xlsx)；
3. [总体架构](04_技术设计/软件总体架构设计说明书.md)与 ADR；
4. [C++ API](05_接口与数据/C++接口说明.md)、[Python API](05_接口与数据/Python接口说明.md)、[插件 API](05_接口与数据/算法插件接口说明.md)；
5. [测试计划](06_测试与验收/测试计划.md)与[WBS](07_项目计划/工作分解结构WBS.xlsx)。

## 3. 证据等级

- E1：原始需求/术语/评审文档；
- E2：可运行 Web 原型和审计截图，仅证明交互表达；
- E3：本交付生成的结构化追踪、接口、资产和测试数据；
- E4：生产 C++、真实 DSP/GPU/模型和性能结果，当前尚未形成，不得宣称完成。

## 4. 自动化入口

运行 `scripts/validate_delivery.py` 复核文件、链接、编号、架构、API、Excel、SVG、PNG、许可证和追踪；依赖获取见 `scripts/fetch_dependencies.ps1`。""",
       "全阶段", "全部输入材料", "建立完整交付导航与证据边界", common_refs, ["生产实现与 E4 证据待后续里程碑形成。"])

    md("CHANGELOG.md", "SS-DOC-CHG-001", "文档集变更日志",
       """## V1.0.0 · 2026-07-22

- 建立十个公共基础模块、分层依赖、API/ABI/SDK/插件/版本边界；
- 将原 SRS 全部稳定需求映射到统一 FR/NFR 编号并生成追踪矩阵；
- 归档 Web 原型与审计截图，新增可浏览的标准截图目录；
- 生成 20 项 ADR、5 份 Excel、品牌 SVG/PNG/ICO、图标、测试数据和依赖锁；
- 建立自动化校验和诚实的证据/未决事项说明；
- 自动审核全部 Markdown，并发布 BL1.0 正式文档基线。""",
       "配置管理", "本次总任务", "首次完整发布", common_refs)

    md("00_交付说明/版本说明.md", "SS-DEL-VER-001", "版本说明",
       """## 1. 版本基线

| 对象 | 版本 | 兼容承诺 |
|---|---|---|
| 文档集 | 1.0.0 | 同主版本修订保持编号稳定 |
| Signal Platform SDK | 1.0.0-candidate | 评审冻结前可调整；冻结后遵循 SemVer |
| Plugin ABI | 1 | 入口符号和 POD 布局按 ABI 主版本管理 |
| Workspace schema | 1.0 | 读同主版本旧次版本；忽略未知可选字段 |
| Dataset schema | 1.0 | 清单向后兼容；破坏性变更迁移 |
| Signal Studio 应用 | R1 设计基线 | 尚无生产 C++ 交付 |
| Web 原型 | 2026-07-22 审计快照 | 仅模拟交互，不是产品实现 |

## 2. 多应用兼容

每个应用依赖 `signal-platform/<major>` 包集合，通过服务注册和公共数据契约组合能力。应用品牌、菜单和工作流位于 `apps/*`；主题令牌可扩展但公共组件不得硬编码应用名。不同应用可并存安装，共享只读运行库、使用各自配置和项目命名空间。

## 3. 范围

本版覆盖设计、接口、测试和计划基线；不包含真实 C++ 可执行程序、100 GB 数据集或性能通过证据。""",
       "配置管理", "总任务、SRS V2.0", "冻结多层版本口径", common_refs)

    md("00_交付说明/已知问题清单.md", "SS-DEL-ISSUE-001", "已知问题清单",
       """## 问题清单

| 编号 | 级别 | 问题 | 影响 | 处理/关闭条件 |
|---|---|---|---|---|
| ISSUE-001 | P0 | W05 原型文件事实/进度分母硬编码 | 进度与部分读取边界不可信 | ReadPlan 单一事实源；AT-20 通过 |
| ISSUE-002 | P0 | 原型存在跨数据源旧对象显示风险 | 可能误判结果来源 | DataSourceVersionId 外键；AT-21 通过 |
| ISSUE-003 | P1 | Web 原型未执行真实 RAW/FFT/STFT/GPU/模型 | 不能证明数值和性能 | E4 生产测试报告 |
| ISSUE-004 | P1 | 1280×720 部分命中区域小于 28 px | 可用性与高 DPI 风险 | Qt 实现扩大透明命中区；AT-24 通过 |
| ISSUE-005 | P1 | Canvas 文本摘要不足 | 辅助技术无法读取范围/单位/质量 | 等价文本与数据表入口 |
| ISSUE-006 | P1 | FFTW 为 GPL，默认闭源分发不兼容 | 法律风险 | 默认 oneMKL；FFTW 仅 GPL 构建 |
| ISSUE-007 | P1 | CUDA、驱动和 ONNX EP 组合尚无设备矩阵 | GPU 路径不确定 | MS-07 设备矩阵和 CPU 回退 |
| ISSUE-008 | P1 | D1-D4 大规模实体数据不在仓库 | 性能验收无法立即执行 | 受控环境生成/导入并登记 SHA256 |

问题状态为开放不等于交付不完整；表示相应生产证据必须在里程碑关闭，不能被文档虚构。""",
       "评审与验收", "综合评审、原型证据、依赖审计", "显式记录冲突和关闭条件", common_refs)

    md("00_交付说明/原型评审结论.md", "SS-DEL-PR-001", "原型评审结论",
       """## 1. 结论

**有条件通过，可进入平台架构和生产实现；不得将原型标记为产品完成。** 原型已覆盖 7 主页面、5 流程界面、高密度工作台、时间导航、共享频率视口、任务暂停/取消/只读前缀、结果/插件/诊断等关键状态。其证据等级为 E2。

## 2. 开发前必须关闭

1. 冻结 `SignalDescriptor`、范围对象、正交状态机和 `ViewRequestId` 原子提交；
2. 去除文件事实硬编码，统一 ReadPlan；
3. 建立 `DataSourceVersionId` 来源隔离；
4. 确认公共 Workbench 与 Signal Studio 专属页面边界；
5. 以 28 px 命中区、键盘遍历、文本摘要作为 Qt UI 门禁。

## 3. 平台化判断

原型信息架构可作为 SignalWorkbench 输入，但当前 HTML 代码为单文件应用，不构成可复用库证据。生产实现必须先建设公共模块，再实现 Signal Studio 页面编排。""",
       "原型评审", "交互原型、11 张审计截图、综合评审", "给出有条件通过结论与开发门禁", common_refs)

    md("00_交付说明/自动审核批准记录.md", "SS-DEL-APR-001", "自动审核批准记录",
       """## 1. 批准结论

审核批次 `APR-20260722-001` 批准 `Signal-Studio-Dev-Docs/` 下的全部 Markdown，形成 `BL1.0` 正式文档基线。批准对象是文档内容、编号、关系、资产引用和实施边界，不是尚未实现的软件功能或尚未执行的生产验收。

## 2. 审核范围

审核覆盖根目录报告、交付说明、需求、原型设计、用户界面（User Interface，UI）规范、技术设计、接口与数据、测试与验收、项目计划、参考资料和 20 项架构决策记录（Architecture Decision Record，ADR）。机器可读清单、Excel、图像、原型和测试数据作为文档证据一并校验，但不改变其各自的实现或测试状态。

## 3. 自动批准门禁

| 门禁 | 通过条件 | 证据 |
|---|---|---|
| 文档结构 | 元数据、内容类型、参考资料、未决事项、变更记录齐全 | `validation-results.json` |
| 批准状态 | 全部 Markdown 为 `BL1.0 / 已批准` | 自动状态扫描 |
| 内容质量 | 文档非空且内容完整，无占位符、无语言代码围栏、失效相对链接和绝对路径 | 写作与 Markdown 检查 |
| 工程一致性 | 编号唯一、需求追踪 100%、架构无环、公共 API 无第三方类型泄漏 | 结构化目录与校验报告 |
| 交付资产 | Excel、SVG、PNG、HTML、许可证、依赖锁和测试数据通过校验 | `validation-report.md` |

## 4. 写作规则适配

审核采用 Writing Guidelines 的可机械验证规则：声明内容类型；标题描述具体对象；段落和列表保持工程任务导向；代码围栏包含语言；链接使用可识别锚文本；禁用占位表达、无效省略号和破折号标点。英文缩写、引号和 Vercel 专用链接规则不替代本任务规定的简体中文、统一编号、固定文档名与顶部元数据格式。

## 5. 批准边界

文档获批不关闭正文中明确列出的 `待确认`、技术建议、开放风险、生产测试或许可证复核事项。`E4` 生产实现证据、真实数字信号处理（Digital Signal Processing，DSP）、图形处理器（Graphics Processing Unit，GPU）、模型运行、100 GB 数据和性能验收仍由对应里程碑产生。任何后续修改必须更新版本、变更记录、追踪关系并重新运行自动校验。""",
       "文档基线批准", "用户自动审核指令、总任务提示词、自动校验结果、Writing Guidelines", "自动审核全部文档并发布正式基线", common_refs,
       ["生产实现、真实算法、设备矩阵和性能验收不在本次文档批准范围内。"])

    # Requirements.
    srs_body = f"""## 1. 产品与平台目标

Signal Platform 提供信号数据、计算、DSP、任务、可视化、工作台、插件、模型和数据集的可复用能力；Signal Studio 负责离线信号分析工作流和品牌编排。目标应用包括 Signal Generator、Dataset Builder/Manager、Model Trainer/Evaluator、Inference Studio、Headless CLI。

## 2. 角色与场景

角色包括信号分析工程师、算法工程师、测试工程师、交付运维和插件开发者。核心场景为大文件渐进浏览、宽窄带联动、可追溯分析、批量生成/数据集/训练/推理任务，以及多宿主 SDK 复用。

## 3. 核心不变量

1. `LoadedDataRange` 是时间导航的唯一总范围；Selection 不进入纯导航条；
2. PSD 与 STFT 共享 `FrequencyViewport`，频率内部为整数 Hz；
3. 任务状态与数据可用状态正交，取消可产生 `cancelled + partial_ready`；
4. 所有下游对象绑定 `DataSourceVersionId`；旧请求不得覆盖新 `ViewRequestId`；
5. 隐藏时域/PSD 后停止其专属准备、观察与绘制，不停止显式共享缓存任务；
6. 公共 API 不暴露第三方类型，应用不直接调用 oneMKL/cuFFT/ONNX Runtime；
7. 需求建议值不能冒充批准值，原型模拟不能冒充真实 DSP。

## 4. 需求详表

每行都具备来源、层级、优先级、验收、页面、接口、测试、里程碑和复用目标；完整的前置/输入/处理/输出/异常字段见 `requirements.json` 与 Excel。

{requirement_table(reqs)}

## 5. 外部接口与约束

Windows 10/11 x64、Qt 6、CMake/MSVC、离线默认；C++20 为建议基线。源数据只读，项目/缓存/结果分离。量化性能、可靠性、数值、安全、可访问和可维护需求保留自 SRS V2.0，并由测试 ID 跟踪。

## 6. 验收规则

全部 P0 需求必须有自动化或可复现测试并通过；P1 未完成须有批准偏差。阻断、数值错误、来源错配、许可证和 ABI 违规为发布门禁。"""
    md("01_需求/软件需求规格说明书.md", "SS-SRS-PLT-001", "软件需求规格说明书", srs_body,
       "需求至验收", "SRS V2.0、综合评审、术语契约、平台总任务", "映射全量需求并新增平台复用约束", common_refs,
       ["内置算法/解调清单、主渲染后端和插件隔离级别需在 MS-00 冻结。"])
    write_json(ROOT / "01_需求/requirements.json", {"schema": "signal.requirements/1.0", "generated": DATE, "requirements": reqs})

    functions = {}
    for r in reqs:
        functions.setdefault(r["library"], []).append(r)
    fbody = ["## 功能分层", "", "| 基础库/层 | 公共能力数 | 可独立发布 | 需要 SDK | 第三方适配 | 原型覆盖 | 里程碑 |", "|---|---:|---|---|---|---|---|"]
    for lib, rows in functions.items():
        fbody.append(f"| {lib} | {len(rows)} | 是 | {'是' if lib in {'SignalPluginSDK','SignalModelRuntime','SignalDataset'} else '内部开发包'} | 是，全部经 Adapter | {'部分交互' if lib in {'SignalVisualization','SignalWorkbench','SignalTaskRuntime'} else '无生产证据'} | {sorted({r['milestone'] for r in rows})[0]} |")
    fbody += ["", "## 应用专属能力", "", "Signal Studio 专属范围仅包括 P01/P02/W02 的业务编排、应用配置、品牌和菜单。RAW/IQ、FFT/STFT、任务、缓存、插件、模型、图表、单位、日志和项目基础设施均不得复制。", "", "## 后续应用复用路径", "", "Signal Generator 复用 Data/DSP/Task/Visualization；Dataset 工具复用 Data/Dataset/Task；训练评估复用 Dataset/Model/Compute/Task；Inference Studio 复用 Data/DSP/Model/Visualization/Workbench。"]
    md("01_需求/功能清单.md", "SS-REQ-FUNC-001", "功能清单", "\n".join(fbody), "产品规划", "需求基线与平台模块", "区分公共能力和应用薄壳", common_refs)

    nfrs = [r for r in reqs if r["id"].startswith("NFR-")]
    md("01_需求/非功能需求说明.md", "SS-REQ-NFR-001", "非功能需求说明",
       f"""## 1. 质量门禁

性能、可靠性、数值、安全、可用性、模块化、复用、API/ABI 稳定、二进制兼容、多应用共存、依赖隔离、插件兼容、SDK 文档、独立构建测试、SemVer、向后兼容、可观测和部署均为可验收要求。

## 2. 关键量化指标

- UI 命令反馈 P95 ≤ 50 ms；图表连续交互 ≥ 30 FPS；取消确认 P95 ≤ 200 ms；
- 标准浏览内存 ≤ 4 GiB 且不随文件大小线性增长；
- FFT/PSD 与批准参考误差 ≤ 0.1 dB，时间误差 ≤ 1 样本，频率误差 ≤ 0.5 bin；
- 隐藏专属视图 500 ms 内无其专属绘制/计算活动；
- 公共模块独立构建、单测、包消费、API/ABI 兼容和多应用复用测试通过。

## 3. 非功能需求追踪

{requirement_table(nfrs)}""",
       "架构、开发、测试", "SRS V2.0 与平台任务", "补充平台化和兼容性质量属性", common_refs)

    # Prototype.
    page_rows = "\n".join(f"| {p['id']} | {p['code']} {p['name']} | {p['owner']} | {p['reuse']} | {p['components']} |" for p in pages)
    md("02_原型设计/原型设计说明书.md", "SS-UI-PD-001", "原型设计说明书",
       f"""## 1. 信息架构与边界

原型包含 7 个持久主页面和 5 个流程界面。主窗口的菜单、工具栏、Dock、任务/结果/日志、设置与诊断归 SignalWorkbench；P02 的宽带分析编排和 W02 通道创建归 Signal Studio。图表组件归 SignalVisualization。

| 页面ID | 页面 | 归属 | 复用性 | 公共组件 |
|---|---|---|---|---|
{page_rows}

## 2. P02 基线

时间导航范围等于实际读入 `LoadedDataRange`，高度紧凑；时域、PSD、STFT 同窗，PSD/STFT 共享频率轴。滚轮只缩放频率横轴；右键左至右选择，右至左重置；三个谱图可重排和调高；频率单位自适应并保留整数 Hz 真值。时域/PSD 可隐藏，隐藏后停止专属计算和渲染。瀑布支持色阶、参考电平和动态范围；纵轴为时间 ms，标签不遮挡内容。

## 3. 原型证据

`原型源文件/Signal Studio交互原型_基线归档.html` 是确定性模拟；`页面截图/audit/` 是输入审计副本；`页面截图/标准截图/` 由本次真实浏览器渲染。三者都不证明生产 Qt/C++ 或真实 DSP 已完成。

## 4. 响应式与可访问

1600×900 为标准基线，1280×720 为最小基线。紧凑控件视觉尺寸可小，但透明命中区不小于 28×28 逻辑像素。Canvas/图表需可访问名称、范围/单位/质量文本摘要和键盘等价操作。""",
       "原型至详细设计", "原型 V1.0、SRS V2.0、审计截图", "重申公共工作台与应用边界", common_refs)

    md("02_原型设计/交互规格说明书.md", "SS-UI-INT-001", "交互规格说明书",
       """## 1. 公共命令模型

所有菜单、工具栏、快捷键和自动化入口注册为 `CommandId`，由 `CanExecute`、参数 schema、权限和撤销策略约束。页面不得直接连接具体后端。

## 2. 强制交互

| 交互ID | 场景 | 状态变化 | 关键验收 |
|---|---|---|---|
| INT-DAT-001 | 导入 | validate → read-plan → task | 文件事实来自同一 ReadPlan |
| INT-TASK-001 | 暂停/继续 | running ↔ paused | 帧边界和 TaskId 不变 |
| INT-TASK-002 | 取消 | running/paused → cancelling → cancelled | 发布已校验只读前缀 |
| INT-NAV-001 | 时间导航 | 新 CurrentTimeViewport/ViewRequestId | 三图同窗、旧响应不提交 |
| INT-VIS-001 | 频率滚轮 | 仅 FrequencyViewport | PSD/STFT/属性/状态栏同步 |
| INT-VIS-002 | 右键选择/重置 | 正向裁剪、反向恢复 | 1 Hz 真值；显示起止和带宽 |
| INT-VIS-003 | 隐藏图表 | visible → hidden | 500 ms 内停止专属准备和绘制 |
| INT-VIS-004 | 色阶/Ref/DR | 仅显示映射 | 不创建文件读取/时间请求 |
| INT-WB-001 | Task Center | 过滤/定位/控制 | 使用任务运行时单一状态源 |
| INT-WB-002 | Result Center | 按来源过滤/定位 | 不跨 DataSourceVersionId 误归类 |

## 3. 错误恢复

参数错误原地保留输入；解析失败不发布未校验数据；可重试、返回修改、取消或查看日志。缓存/GPU/插件失败降级必须显示实际后端和影响，且不修改源文件。""",
       "交互设计与测试", "当前原型、综合评审", "建立统一交互 ID 和状态契约", common_refs)

    md("02_原型设计/状态与异常场景说明.md", "SS-UI-STATE-001", "状态与异常场景说明",
       """## 1. 正交状态

`ReadTaskState` 与 `DataSourceReadState` 独立：`cancelled + partial_ready` 是合法组合。`OverviewQuality`、`ResultValidity`、`ViewRequestQuality` 也不得复用任务状态枚举。

```mermaid
stateDiagram-v2
  [*] --> queued
  queued --> running
  running --> paused
  paused --> running
  running --> cancelling
  paused --> cancelling
  cancelling --> cancelled
  running --> completed
  running --> failed
```

## 2. 异常目录

| 异常ID | 条件 | 用户语义 | 数据处理 | 恢复 |
|---|---|---|---|---|
| EX-DAT-001 | 文件尾非完整帧 | 样本帧对齐失败 | 不发布未校验尾部 | 修改格式或使用完整前缀 |
| EX-DAT-002 | 源文件变化 | 数据源版本过期 | 缓存/结果标 stale | 重验证或新建版本 |
| EX-CACHE-001 | 磁盘满/瓦片损坏 | 缓存降级 | 源文件与项目不变 | 清理/换目录/按需计算 |
| EX-COMPUTE-001 | GPU 不可用 | 回退 CPU | 记录实际后端 | 重新探测设备 |
| EX-PLG-001 | 插件崩溃/不兼容 | 隔离并拒绝加载 | 不提交结果 | 安全模式/升级插件 |
| EX-MODEL-001 | 模型输入不兼容 | 不适用 | 不运行推理 | 选择兼容模型/前处理 |
| EX-PROV-001 | 来源键不一致 | 阻止分析 | 不显示为当前对象 | 刷新上下文/诊断 |
| EX-TASK-001 | 依赖失败/取消 | 依赖失败 | 不发布半成品 | 修复依赖后重试 |

所有异常映射稳定 ERR 码，含用户消息、技术详情、建议、TaskId、DataSourceVersionId。""",
       "详细设计与测试", "术语契约、综合评审", "统一正交状态与异常目录", common_refs)

    # UI specs.
    md("03_UI规范/UI视觉规范.md", "SS-UI-DS-001", "UI 视觉规范",
       """## 1. Signal Design System

Design System 分为 Signal Design Tokens、Signal UI Components、Signal Chart Components、Signal Workbench Components 和 Signal Icon System。公共主题不含应用品牌；Signal Studio 仅在应用层注入 Logo、名称和强调色。

## 2. 令牌

| 类别 | 令牌 | 建议值 | 规则 |
|---|---|---|---|
| 背景 | `surface.canvas` | `#08111F` | 图表/工作台主背景 |
| 面板 | `surface.panel` | `#0E1B2D` | Dock 与卡片 |
| 文本 | `text.primary` | `#E5F1FF` | 与背景对比 ≥ 4.5:1 |
| 辅助 | `text.muted` | `#8FA8C2` | 仅辅助信息 |
| 强调 | `accent.cyan` | `#20D3EE` | 当前/交互，不独立表达状态 |
| 危险 | `status.error` | `#FF6B7A` | 同时配文字/图标 |
| 间距 | `space.1..6` | 4/8/12/16/24/32 px | 4 px 基线 |
| 字体 | `font.ui` | Microsoft YaHei UI | 中文优先；等宽数值另用 Cascadia Mono |

## 3. 密度与可访问

正文 12–13 px、辅助 11–12 px；视觉紧凑不降低 28×28 逻辑像素命中区。焦点环 2 px；状态不只依赖颜色；图表提供文本摘要。100%–200% DPI 使用逻辑尺寸和矢量图标。

## 4. 品牌隔离

Signal Studio 品牌资产为应用层专有；公共库、SDK、主题和工作台组件不得包含 `Signal Studio` 字样、SS 图形或默认应用菜单。其他工具使用独立 Logo。""",
       "UI 实现", "原型视觉令牌、可访问性审计", "建立公共 Design System 和品牌隔离", common_refs)

    md("03_UI规范/图表显示规范.md", "SS-UI-CHART-001", "图表显示规范",
       """## 1. 公共组件

`TimeWaveformView`、`TimeNavigator`、`SpectrumView`、`PSDView`、`WaterfallView`、`SpectrogramView`、`ConstellationView`、`EyeDiagramView`、`FrequencyAxis`、`TimeAxis`、`FrequencyViewport`、`TimeViewport`、`MeasurementOverlay`、`SelectionOverlay` 全部归 SignalVisualization。

## 2. 坐标与数据

- 时间真值为 64 位样本索引半开区间，秒仅显示；
- 频率真值为整数 Hz，单位按范围选择 Hz/kHz/MHz/GHz 且保留 1 Hz；
- PSD 纵轴为 dB/Hz，必须记录窗增益、ENBW、FFT 帧数和参考量；
- 瀑布/时频图横轴频率、纵轴时间 ms 向下递增；
- PSD 和 STFT 共享频率视口，所有同窗图使用同一 ViewRequestId。

## 3. 交互与调度

滚轮只缩放频率横轴；指针锚定。右键左至右框选、右至左重置。三个谱图可调高、重排且持久化。隐藏组件必须断开其观察器、绘图和专属任务；共享缓存只有在显式任务生命周期下继续。

## 4. 渲染后端

公共接口只接收平台数据/视口/图层契约。Qt/Qwt/OpenGL/RHI 仅在私有适配器。热路径禁止每帧分配；纹理/瓦片按资源预算复用；CPU/GPU 渲染必须具备可比截图或数值采样验证。""",
       "UI/DSP 联合实现", "SRS 图谱与导航需求", "将图表定义为跨应用组件", common_refs)

    md("03_UI规范/时频图色阶规范.md", "SS-UI-CMAP-001", "时频图色阶规范",
       """## 1. 数据到颜色

显示值 $v$、参考电平 $R$、动态范围 $D$ 的归一化为 $t=\\operatorname{clamp}((v-(R-D))/D,0,1)$。色阶只改变 $t\\to color$ 的显示映射，不修改 PSD/STFT 标量、不创建读取任务或时间请求。

## 2. 预设

默认建议使用感知均匀 Viridis；Industrial 为工程主题；Turbo 只用于高对比诊断；Grayscale 用于打印。每个预设提供色盲检查、灰度单调性说明、NaN/低于量程/高于量程专色。

## 3. UI

色标显示单位、Ref、DR、最小/最大和色图名。标签背景透明，必要时文字描边或外置轴。键盘可调整 Ref/DR；恢复默认不触发数据请求。截图必须保留色标和参数摘要，除非用户显式关闭。""",
       "可视化实现与验收", "SRS VIS-013/016、原型", "冻结色阶映射与非计算语义", common_refs)

    # Technical design.
    arch_body = """## 1. 架构结论

采用 **Monorepo + CMake targets + vcpkg manifest + 混合链接**。十个公共模块先于应用建设；Signal Studio 位于 `apps/signal-studio`，只编排公开服务。普通 C++ 库以隐藏符号/PIMPL 的共享库发布，性能敏感模板放开发包；插件边界使用版本化 C ABI，C++ wrapper 仅为便利层。

```text
apps/{signal-studio,signal-generator,dataset-manager,model-trainer,inference-studio}
libs/{signal-core,signal-data,signal-dsp,signal-compute,signal-task-runtime,
      signal-visualization,signal-workbench,signal-plugin-sdk,signal-model-runtime,signal-dataset}
plugins/{importers,exporters,algorithms,models,visualizations}
tests/{unit,integration,performance,compatibility,golden-data}
```

## 2. 分层与依赖

```mermaid
flowchart TB
  Apps["Applications: Studio / Generator / Dataset / Model / Inference"] --> WB[SignalWorkbench]
  Apps --> VIS[SignalVisualization]
  Apps --> MR[SignalModelRuntime]
  Apps --> DS[SignalDataset]
  Apps --> DSP[SignalDSP]
  WB --> VIS
  WB --> TASK[SignalTaskRuntime]
  VIS --> DATA[SignalData]
  VIS --> TASK
  MR --> DATA
  MR --> TASK
  MR --> CMP[SignalCompute]
  DS --> DATA
  DS --> TASK
  DSP --> DATA
  DSP --> CMP
  TASK --> CMP
  DATA --> CORE[SignalCore]
  CMP --> CORE
  PLG[SignalPluginSDK] --> DATA
  PLG --> TASK
  PLG --> CORE
```

允许方向与 `architecture.json` 一致。Core 不依赖 Qt Widgets；DSP/Data 不依赖页面；UI 不直接操作文件句柄；应用不调用 oneMKL/cuFFT/ONNX Runtime。CI 对 DAG 做循环检测，并扫描公共签名的第三方类型。

## 3. CMake 与包

公开 targets：`Signal::Core`、`Signal::Data`、`Signal::Compute`、`Signal::DSP`、`Signal::TaskRuntime`、`Signal::Visualization`、`Signal::Workbench`、`Signal::PluginSDK`、`Signal::ModelRuntime`、`Signal::Dataset`。每个包包含 runtime/dev/debug symbols/licenses/SBOM；包消费者测试使用 `find_package(SignalPlatform 1 CONFIG REQUIRED)`。

## 4. API、ABI 与版本

公共头文件位于 `include/signal/<module>`，私有实现位于 `src`。导出类用 PIMPL，异常不跨 ABI；插件以 `signal_plugin_query_v1` 交换 POD 与函数表。SDK、schema、插件 ABI、应用独立 SemVer；弃用至少跨一个次版本并在编译/运行日志提示。

## 5. 第三方适配

FFT 默认 oneMKL，CUDA/cuFFT 为可选后端；FFTW 仅许可兼容构建。Eigen、TBB、HDF5、ONNX Runtime、spdlog、fmt、GoogleTest、Google Benchmark 仅出现在私有适配器/测试中。上层只见平台类型。Qt 图形方案需经许可证和百万点/时频瓦片基准后冻结。

## 6. 测试边界

每个模块独立单测；相邻模块契约测试；适配器黄金数据和 CPU/GPU 一致性；包消费/ABI/API/插件兼容；第二宿主复用；应用 E2E。公共测试禁止通过 Signal Studio 私有 UI 驱动。"""
    md("04_技术设计/软件总体架构设计说明书.md", "SS-ARCH-001", "软件总体架构设计说明书", arch_body,
       "架构至发布", "平台总任务、SRS、术语契约", "建立可复用平台为最高优先级", common_refs,
       ["Qt 图表实现、插件进程隔离默认级别和二进制签名方案待 MS-00 基准/安全评审冻结。"])
    write_json(ROOT / "04_技术设计/architecture.json", architecture())

    module_sections = []
    caps_by_lib = {}
    for c in libs:
        caps_by_lib.setdefault(c["library"], []).append(c)
    for node in architecture()["nodes"]:
        lib = node["name"]
        cap_text = "、".join(f"{c['id']} {c['capability']}" for c in caps_by_lib[lib])
        module_sections.append(f"""## {lib}

- 责任：{cap_text}。
- 公共 API：{', '.join(a['id'] for a in apis if a['module'] == lib) or '通过相邻公共服务契约'}；命名空间 `{node['public_namespace']}`。
- 数据/线程：不可变请求优先；UI 仅主线程，IO/计算/推理均由 TaskRuntime 调度；只读视图明确生命周期。
- 依赖：{', '.join(node['deps']) or '无公共模块依赖'}；第三方仅私有 Adapter。
- 错误/性能：统一 `Result<Error>`、追踪 ID、资源预算和基准；不得静默回退。
- 测试：单元、契约、适配器、性能、包消费；复用目标：{APP_TARGETS[lib]}。
- 发布：`{lib.lower()}-runtime/dev`，SemVer 1.x；非目标：应用页面、品牌和应用专属工作流。
""")
    md("04_技术设计/模块详细设计说明书.md", "SS-DES-MOD-001", "模块详细设计说明书",
       "## 设计总则\n\n模块按公开契约、私有实现、适配器和测试夹具四区组织；下列每个模块均可独立构建、安装和测试。\n\n" + "\n".join(module_sections),
       "详细设计与实现", "总体架构、API 目录", "定义十模块责任、API、线程、依赖和非目标", common_refs)

    md("04_技术设计/线程与任务调度设计.md", "SS-DES-TASK-001", "线程与任务调度设计",
       """## 1. 统一任务契约

`TaskSpec` 包含 TaskId、TaskType、Priority、ResourceProfile、ProgressModel、CancellationToken、PauseToken、Dependencies、Artifacts、ResultPolicy、ErrorPolicy、Logs、Metrics、IdempotencyKey 和 provenance。进入 queued 后参数冻结。

## 2. 调度域

| 域 | 工作 | 并发/取消 |
|---|---|---|
| UI | Qt 对象与原子提交 | 单线程；不得阻塞 >50 ms |
| IO | 分块读取、缓存、导出 | 按设备限流；块边界取消 |
| CPU | FFT/STFT/滤波/生成/预处理 | TBB 预算；保留 UI 核 |
| GPU | FFT、推理、纹理 | 设备队列、stream、显存配额 |
| 外部进程 | Python/不可信插件/训练 | 进程监督、IPC、硬超时 |

支持加载、FFT/STFT、信号生成、批量样本、数据集导入、训练、评估和推理。优先级为 Interactive > Foreground > Background；同资源池使用加权公平与配额，禁止训练饿死交互。

## 3. 状态与恢复

queued/running/paused/cancelling/cancelled/completed/failed/dependency_failed/stale。暂停点仅在声明安全的块边界；取消幂等，P95 200 ms 确认。持久任务写 append-only journal，制品先临时写后校验/原子提交；崩溃恢复重新验证而非假定成功。

## 4. 跨工具复用

Task Center 只订阅公共 `TaskEvent`；应用提供 TaskType 显示适配，不建立私有队列。Headless CLI 使用同一服务和 journal。测试包含优先级倒置、DAG 传播、暂停边界、取消时延、崩溃恢复和多应用并存。""",
       "任务运行时实现", "SRS TSK、W05、平台任务", "将所有工具统一到跨应用 TaskRuntime", common_refs)

    md("04_技术设计/大文件与缓存设计.md", "SS-DES-CACHE-001", "大文件与缓存设计",
       """## 1. 适用对象

同一缓存基础设施服务 RAW/IQ、仿真信号、样本集、时域金字塔、PSD/STFT 瓦片、模型输入、训练/推理批次。源只读；项目元数据和已提交结果不得被 LRU 当作缓存删除。

## 2. 数据通路

`IDataSource` 分块读取 → 帧对齐验证 → `SampleBlockView` → 多分辨率/算法任务 → 临时制品 → SHA256/元数据 → 原子提交。内存映射仅封装在 SignalData；UI 获得租约视图，不持有文件句柄。

缓存键包含 SourceFingerprint、DataSourceVersionId、LoadedDataRange、算法/依赖版本、参数摘要、时间/频率视口、像素尺寸和质量。未提交瓦片无有效标记。部分读取的索引上界严格等于 LoadedDataRange.end。

## 3. 层级与预算

L0 小块/热点内存、L1 项目 NVMe 瓦片、L2 可选共享只读制品。默认内存 25%（10%–60%），预留系统 4 GiB；磁盘按项目 pin + 全局 LRU。预取低优先并可取消。

## 4. 一致性与恢复

使用 manifest + content digest；源/描述符/算法变化只失效相关键。磁盘满、权限或损坏时降级按需计算并记录。跨进程锁带租约/持有者；崩溃后清理无提交标记的临时对象。""",
       "数据与缓存实现", "SRS DAT/IDX、术语契约", "扩展到生成、数据集、训练与推理缓存", common_refs)

    md("04_技术设计/GPU加速设计.md", "SS-DES-GPU-001", "GPU 加速设计",
       """## 1. 边界

SignalCompute 统一 GPU Runtime，服务 FFT/STFT、批量生成、预处理、模型训练辅助、推理和可视化纹理。应用层只请求 `Workload`，不得调用 CUDA/cuFFT/ONNX EP。

## 2. 选择与降级

`Auto` 根据能力、数据规模、队列、传输成本和显存选择 CPU/CPU_SIMD/CPU_Multithread/CUDA。选择记录设备、驱动、工具链、后端版本和原因。OOM、驱动或结果校验失败回退 CPU，并在 UI/结果 provenance 明示。

## 3. 资源与并发

每设备管理 context、stream 池、计划缓存、pinned/device 内存池和配额。交互预留高优先 stream；训练/批量任务可被限流但不强制中断不安全 kernel。跨应用默认各进程配额，未来守护进程模式为待确认。

## 4. 正确性与性能

以已批准 CPU 参考验证幅值/相位/PSD/STFT，阈值按 dtype/算法冻结。测试小/大/质数/2 的幂 FFT、非连续块、取消、OOM、多 stream 和设备切换；报告端到端吞吐，不只 kernel 时间。CUDA 12.8.1 为可选锁定工具链，安装需 EULA，默认包不强制依赖。""",
       "计算与部署", "平台任务、依赖锁、SRS 性能", "建立跨工具统一 GPU Runtime", common_refs)

    # ADRs 001-020.
    adr_topics = [
        (1, "采用 Monorepo 与统一 CMake 超级构建", "所有公共库、应用、插件模板和契约测试在同仓演进；发布时按包拆分。", "多仓库会增加早期跨库变更和原子兼容测试成本。"),
        (2, "公共库采用混合链接与隐藏实现", "运行库以共享库/PIMPL 为主，模板和小型值类型在开发包；插件使用 C ABI。", "全静态放大安装且许可证复杂；全头文件破坏 ABI。"),
        (3, "Qt 6 仅进入 Workbench 与 Visualization 私有实现", "Core/Data/DSP/Compute/Task 不依赖 Qt Widgets；UI 经服务接口消费。", "全栈 Qt 类型会污染 Headless/Python/SDK。"),
        (4, "不可变数据版本与来源链", "DataSourceVersion、TaskSpec、Result 发布后不可原位修改；全部带哈希和版本。", "可变共享对象难以复现并造成跨源污染。"),
        (5, "vcpkg manifest 与固定 baseline", "使用固定 commit/版本和离线缓存；外部 CUDA 单独 URL+SHA256。", "浮动 FetchContent 不能复现；手工依赖难审计。"),
        (6, "FFT 与矩阵后端适配", "oneMKL 为默认 CPU，cuFFT 可选；FFTW 仅许可兼容构建；Eigen 不出公共 API。", "自研 FFT/矩阵禁止；直接调用阻碍替换。"),
        (7, "统一任务运行时", "所有 IO/DSP/生成/数据集/模型任务使用 SignalTaskRuntime。", "每个应用私有队列会复制状态、取消和恢复逻辑。"),
        (8, "大文件分块与多分辨率瓦片", "零拷贝分块、内存映射适配、分层缓存和原子制品。", "整文件入内存不满足 10–100 GB。"),
        (9, "计算后端自动选择与显式降级", "CPU/SIMD/TBB/CUDA 经 SignalCompute，provenance 记录实际后端。", "GPU 硬绑定会使兼容路径缺失。"),
        (10, "插件版本化 C ABI 与隔离", "单一 C ABI 入口、POD 句柄、manifest、权限和可选进程隔离。", "Qt/C++ 私有对象跨边界导致 ABI 脆弱。"),
        (11, "ONNX Runtime 为默认推理后端", "ModelRuntime 优先 ORT CPU/CUDA；LibTorch 作为延后适配器。", "同时默认两套运行时增加体积与冲突。"),
        (12, "数据集清单与可插拔分片格式", "公共 manifest/index；HDF5 首个结构化适配，WebDataset 用于流式训练。", "单一物理格式不能覆盖管理与训练。"),
        (13, "基础能力库拆分策略", "按 Core/Data/DSP/Compute/Task/Vis/Workbench/Plugin/Model/Dataset 十模块拆分。", "按页面拆分会复制底层能力。"),
        (14, "多应用工作台复用策略", "Workbench 提供壳、命令、Dock、中心和设置；应用注入品牌/工作流。", "复制主框架会造成多应用体验和修复分叉。"),
        (15, "公共数据契约与版本兼容", "schema SemVer、未知可选字段忽略、必填/语义变更升主版本并迁移。", "无版本 JSON 无法安全演进。"),
        (16, "公共可视化组件复用策略", "图表/轴/视口/覆盖层在 SignalVisualization；数据提供者接口解耦。", "页面自绘会重复交互、单位和调度。"),
        (17, "任务运行时跨工具复用", "TaskSpec/TaskEvent/journal 统一，UI/CLI/多应用共享。", "Signal Studio 专属任务类阻碍生成/训练/推理。"),
        (18, "模型与数据集能力边界", "Dataset 管理可复现数据，ModelRuntime 管理模型与推理；训练编排在应用/服务层。", "将训练或数据索引塞入模型运行时会耦合。"),
        (19, "单仓库与多仓库策略", "1.x 使用 Monorepo；稳定后可按发布包镜像，不拆权威来源。", "过早多仓库降低原子变更和兼容门禁。"),
        (20, "SDK 二进制兼容与发布策略", "MSVC x64 ABI 矩阵、C ABI 插件、PIMPL、符号清单、SemVer 和弃用窗口。", "只承诺源码兼容不足以支撑第三方插件。"),
    ]
    adr_dir = ROOT / "04_技术设计/架构决策记录"
    for num, title, decision, alternative in adr_topics:
        body = f"""## 背景

Signal Platform 必须被多个桌面应用、Headless CLI、插件和 Python SDK 复用，并保持可重复构建、许可证合规和可测试边界。

## 决策

{decision}

## 备选方案

{alternative}

## 正面影响

依赖方向可机械校验；公共能力可独立发布和测试；应用保持薄壳；后端和第三方库可替换。

## 代价与风险

需要维护契约、适配器、包和兼容矩阵；短期工作量高于单应用原型，但避免后续被动重构。

## 验证方式

构建 DAG/循环检查、公共头第三方类型扫描、模块单测、包消费、ABI/API、第二宿主和许可证测试。

## 替代关系

若后续 ADR 替代本决策，必须列出迁移、版本影响和失效测试；当前无替代。"""
        md(f"04_技术设计/架构决策记录/ADR-{num:03d}_{title}.md", f"ADR-{num:03d}", f"ADR-{num:03d} {title}", body,
           "架构至发布", "总体架构、平台任务、依赖审计", "记录已采用并自动审核批准的基线决策", common_refs,
           ["本决策已进入 BL1.0；若后续 ADR 替代，必须提供迁移、兼容和验证证据。"])

    # Interfaces and data.
    md("05_接口与数据/RAW-IQ数据格式说明.md", "SS-DATA-RAW-001", "RAW-IQ 数据格式说明",
       """## 1. 适用范围

RAW 是无自描述字节流；IQ/QI 是复信号分量排列，不是全部信号的统称。任何导入必须有 `SignalDescriptor`，不得凭文件名静默猜测影响数值解释的字段。

## 2. 描述符

必填：signalKind、scalarType、componentLayout/order、endianness、sampleRateHz、byteOffset、requestedSampleRange、amplitudeMode/scale。中心频率可为空；多字节需要字节序；单字节为 not_applicable。帧字节数为分量数乘标量字节数，所有范围按完整帧对齐并用 64 位半开样本区间。

```json
{"schema":"signal.raw-descriptor/1.0","signalKind":"complex","scalarType":"int16","componentLayout":"interleaved","componentOrder":"IQ","endianness":"little","sampleRateHz":50000000,"centerFrequencyHz":1245000000,"byteOffset":0,"requestedSampleRange":{"start":0,"end":249693612},"amplitudeMode":"int16_scaled","scaleFactor":0.000030517578125}
```

## 3. 校验与读取

验证文件事实、剩余字节、帧对齐、NaN/Inf/削顶/直流/全零。ReadPlan 的目标为 `min(configuredInitialBytes, remainingFrameAlignedBytes)`。暂停保持边界；取消发布最后完整帧前缀。源始终只读，侧车与缓存不写回源。

## 4. 频率/PSD

实信号默认单边 `0..Fs/2`，复信号默认 `[-Fs/2,Fs/2)` 并可加中心频率。PSD dB/Hz 必须记录窗、ENBW、归一化和参考量；未经标定不得显示 dBm。""",
       "数据实现与测试", "术语契约、SRS DAT", "冻结 RAW/实复/IQ、范围和校验语义", common_refs)

    md("05_接口与数据/工程文件格式说明.md", "SS-DATA-WS-001", "工程文件格式说明",
       """## 1. 格式族

公共 Workspace 保存身份、版本、资源 URI、任务/结果索引、布局引用和扩展命名空间；Signal Studio、Dataset、Training、Inference 项目在 `extensions.<app-id>` 下保存专属字段。公共部分不含应用品牌。

```json
{"schema":"signal.workspace/1.0","workspaceId":"uuid","createdAt":"ISO-8601","resources":[],"tasks":[],"results":[],"extensions":{"org.signalplatform.signal-studio":{"layout":{},"analysisChannels":[]}}}
```

## 2. 版本和未知字段

消费者读取同主版本的旧次版本并忽略未知可选字段；未知必需 capability 导致结构化不兼容错误；破坏性变更升主版本并提供显式迁移。插件扩展使用反向域名命名空间。

## 3. 持久化

UTF-8、规范 JSON、相对 URI 优先；保存到同目录临时文件，fsync、校验后原子替换。源文件、缓存、不可变结果分离。恢复点带父版本和 digest；加载先验证 schema，再解析扩展。

## 4. 安全

拒绝路径穿越、设备路径和未经许可外联 URI；日志脱敏路径；插件不能写未知命名空间。JSON Schema 位于 `schemas/`。""",
       "数据与配置", "SRS PRJ、公共数据契约", "区分公共 Workspace 与应用扩展", common_refs)

    cxx_code = """```cpp
namespace signal::task {
struct TaskSpec final {
  core::Uuid task_id;
  core::String task_type;
  ResourceProfile resources;
  core::Vector<core::Uuid> dependencies;
  core::Hash256 idempotency_key;
};

class SIGNAL_TASK_API ITaskService {
public:
  virtual ~ITaskService() noexcept = default;
  virtual core::Result<TaskHandle> submit(const TaskSpec&) noexcept = 0;
};
}

extern "C" SIGNAL_PLUGIN_EXPORT int
signal_plugin_query_v1(const SignalHostApiV1*, SignalPluginApiV1*) noexcept;
```"""
    md("05_接口与数据/C++接口说明.md", "SS-API-CPP-001", "C++ 接口说明",
       f"""## 1. 命名空间与目标

`signal::core/data/dsp/compute/task/visualization/workbench/plugin/model/dataset` 分别由同名 CMake target 导出。公共 API 用平台值类型、PIMPL/抽象接口和 `Result<T>`；禁止 Qt、Eigen、oneMKL、FFTW、cuFFT、TBB、ONNX Runtime 类型进入签名。

## 2. API 目录

{api_table([a for a in apis if not a['id'].startswith('API-PY')])}

## 3. 示例

{cxx_code}

## 4. 兼容和发布

SDK `MAJOR.MINOR.PATCH`；Plugin ABI 单独整数主版本。Windows ABI 矩阵锁 MSVC toolset/x64/runtime；异常、STL 容器所有权和分配器不得跨插件 ABI。弃用至少保留一个次版本，提供替代 API 和编译警告。包导出 `SignalPlatformConfig.cmake`、targets、headers、symbols、licenses、SBOM 和 examples。

## 5. 线程与生命周期

接口标注 ThreadSafe/UIThread/TaskThread；所有 handle 有明确所有权，回调允许注销并防止退出后调用。耗时调用返回 TaskHandle，不在 UI 线程同步执行。""",
       "SDK 实现", "总体架构、模块设计", "定义公共 API/SDK/ABI 与第三方隔离", common_refs)
    write_json(ROOT / "05_接口与数据/api-catalog.json", {"schema": "signal.api/1.0", "generated": DATE, "apis": apis})

    md("05_接口与数据/Python接口说明.md", "SS-API-PY-001", "Python 接口说明",
       """## 1. 包与范围

包名 `signal_platform`，首版提供 data、dsp、task、model、dataset 子模块；训练框架保留在应用层。pybind11 私有封装 C++ SDK，Python 侧不暴露 Qt 对象。

```python
import signal_platform as sp
source = sp.open_signal("capture.sc16", descriptor)
view = source.read(samples=(0, 1_000_000))   # NumPy 只读/可控零拷贝
task = sp.submit_task(sp.PsdRequest(source=view, fft_size=8192))
result = task.result(timeout=30)
dataset = sp.load_dataset("dataset://demo@1.0")
pred = sp.infer("model://classifier@2.1", dataset.batch(32))
```

## 2. 内存与互操作

NumPy 通过 buffer protocol 交换实/复数组，shape/stride/dtype/endianness 明确；生命周期由 owning capsule 保证。PyTorch 通过 DLPack 为 Preview，设备/stream 必须显式。不得把可写 NumPy 视图指向只读源映射。

## 3. 任务与错误

同步方法仅限轻量元数据；IO/DSP/推理返回 awaitable TaskHandle，支持 status/progress/pause/resume/cancel/logs。C++ Error 映射稳定 Python 异常类并保留 code/details/recovery。

## 4. Wheel 与兼容

支持 CPython 3.11–3.13 x64（建议基线），wheel 名含平台 ABI；每个 wheel 在干净环境运行导入、NumPy、任务、数据集和推理 smoke test。Python API 遵循 SemVer，Preview 能力显式标注。""",
       "Python SDK", "C++ API、平台任务", "定义 Python/NumPy/PyTorch/任务/模型复用", common_refs,
       ["CPython 精确支持矩阵需与企业运行环境在 MS-00 确认。"])

    md("05_接口与数据/算法插件接口说明.md", "SS-API-PLG-001", "算法插件接口说明",
       """## 1. 宿主无关原则

插件不得包含或访问 Signal Studio 页面、MainWindow、Dock 或私有模型。输入输出只使用公共 Data/Task/Result 契约，因而可运行于 Signal Studio、Signal Generator、Dataset Manager、Model Trainer、Inference Studio、Headless CLI 和 Automated Test Runner。

## 2. Manifest

字段：id、version、pluginAbi、sdkRange、architecture、capabilities、input/output contracts、threadModel、permissions、dependencies、license、publisher、signature、contentHashes。未知必需 capability 拒绝加载。

## 3. 生命周期

`query → validate → load → activate → quiesce → deactivate → unload`。入口为 `signal_plugin_query_v1`，交换 host/plugin 函数表和不透明句柄。算法 `describe` 声明实/复、dtype、采样率、长度、设备与参数 schema；`run` 只能通过宿主 Task/Artifact/Log 服务。

## 4. 线程、权限与失败

默认无 UI、无网络、只读输入、受控制品目录。耗时工作必须任务化并响应取消。崩溃/超时/越权导致隔离和稳定错误；不发布半成品。高风险 Python/未知发布者建议进程外。

## 5. SDK

提供 C ABI 头、C++ wrapper、CMake package、算法/格式示例、manifest schema、签名工具、契约测试和兼容检查器。插件包必须携带许可证和 SBOM。""",
       "插件 SDK", "SRS PLG、ADR-010/020", "定义多宿主、无 UI 依赖插件 ABI", common_refs)

    error_rows = []
    modules = ["CORE", "DATA", "DSP", "COMPUTE", "TASK", "VIS", "WB", "PLG", "MODEL", "DSET"]
    descs = ["无效参数或契约", "资源/数据不可用", "取消或过期", "内部适配器失败"]
    for m in modules:
        for i, d in enumerate(descs, 1):
            error_rows.append(f"| ERR-{m}-{i:03d} | SS-{m}-E{i:03d} | {d} | 不修改源；不提交半成品 | 检查详情、修正输入或选择回退 |")
    md("05_接口与数据/错误码说明.md", "SS-API-ERR-001", "错误码说明",
       """## 1. 模型

错误包含 code、category、severity、userMessage、technicalDetails、recoveryActions、retryable、TaskId、ObjectId、DataSourceVersionId、causeChain 和 metricsRef。用户文案可本地化，稳定码不可本地化。

| 编号 | 稳定码 | 语义 | 数据保证 | 恢复 |
|---|---|---|---|---|
""" + "\n".join(error_rows) + """

## 2. 传播

适配器错误在模块边界映射为公共码并保留私有诊断；异常不得越过 C ABI。任务失败与数据状态独立；取消不是普通失败。日志默认脱敏路径和样本，不包含密钥。""",
       "全实现与运维", "统一错误模型、模块设计", "建立十模块稳定错误码", common_refs)

    # JSON schemas.
    schema_dir = ROOT / "05_接口与数据/schemas"
    workspace_schema = {"$schema": "https://json-schema.org/draft/2020-12/schema", "$id": "https://signal-platform.invalid/schema/workspace-1.0.json", "title": "Signal Platform Workspace", "type": "object", "required": ["schema", "workspaceId", "resources", "extensions"], "properties": {"schema": {"const": "signal.workspace/1.0"}, "workspaceId": {"type": "string", "minLength": 1}, "resources": {"type": "array"}, "tasks": {"type": "array"}, "results": {"type": "array"}, "extensions": {"type": "object"}}, "additionalProperties": True}
    plugin_schema = {"$schema": "https://json-schema.org/draft/2020-12/schema", "$id": "https://signal-platform.invalid/schema/plugin-manifest-1.0.json", "title": "Signal Plugin Manifest", "type": "object", "required": ["id", "version", "pluginAbi", "capabilities", "license", "contentHashes"], "properties": {"id": {"type": "string", "pattern": "^[a-z0-9]+([.-][a-z0-9]+)+$"}, "version": {"type": "string"}, "pluginAbi": {"const": 1}, "capabilities": {"type": "array", "items": {"type": "string"}}, "license": {"type": "string"}, "contentHashes": {"type": "object"}}, "additionalProperties": True}
    write_json(schema_dir / "signal-workspace-1.0.schema.json", workspace_schema)
    write_json(schema_dir / "signal-plugin-manifest-1.0.schema.json", plugin_schema)

    # Test and acceptance.
    md("06_测试与验收/测试计划.md", "SS-TEST-PLAN-001", "测试计划",
       """## 1. 策略

测试金字塔：模块单元 → 公共契约/适配器 → 相邻模块集成 → 包消费/API/ABI/插件兼容 → 多应用复用 → Signal Studio E2E → 性能耐久。原型截图回归只验证交互，不替代数值测试。

## 2. 测试域

| 域 | 范围 | 环境/证据 |
|---|---|---|
| 基础库 | 十模块独立单测、错误、边界、资源 | Headless，JUnit/覆盖率 |
| SDK | CMake 消费、wheel、示例编译 | 干净 VM/venv |
| API/ABI | 符号、头文件、布局、弃用 | 1.x 版本对比 |
| 插件 | 多宿主、版本拒绝、崩溃、权限 | Test Runner + 安全模式 |
| 多应用 | Studio + Generator 薄壳并存 | 同平台包、独立配置 |
| 算法 | oneMKL/cuFFT/参考黄金 | 输入 SHA256、误差报告 |
| 大文件 | D1-D4/部分读取/缓存/磁盘满 | 受控数据与 B1 硬件 |
| UI | 1600×900、1366×768、1280×720、100–200% DPI | 截图、键盘、可访问树 |

## 3. 数据与环境

本交付提供 D0 smoke 数据与生成脚本。D1-D3 大数据必须在有容量的验收环境显式生成并登记哈希；D4 未随输入仓库提供，不得伪造。B1 环境记录 CPU、RAM、NVMe、GPU/驱动、后端和构建 ID。

## 4. 门禁

P0 需求测试设计覆盖 100%，发布前实际通过率 100%；阻断/数值/来源错配/许可证/ABI 问题 0。失败用例关联缺陷和重测证据。测试 Excel 的“已设计”不代表执行通过。""",
       "测试设计至发布", "SRS、接口、风险", "覆盖基础库、SDK、插件和多应用复用", common_refs)

    md("06_测试与验收/性能测试方案.md", "SS-TEST-PERF-001", "性能测试方案",
       """## 1. 公共基准

| 基准 | 尺寸/变量 | 指标 |
|---|---|---|
| FFT | 1K–1M、实/复、批量、冷热计划 | samples/s、P50/P95、计划时间、误差 |
| STFT | FFT/overlap/视窗/后端 | frames/s、首结果、显存/内存 |
| 重采样 | 有理比、tap、块大小 | MS/s、延迟、阻带误差 |
| 数据读取 | 顺序/随机/mmap/分块 | GiB/s、P95、CPU、句柄 |
| 缓存 | 热/冷/损坏/磁盘满 | 命中、首屏、恢复 |
| 调度 | 1–1000任务、DAG、取消 | 排队、取消P95、公平性 |
| GPU | H2D/D2H、plan/stream、OOM | 端到端加速比、回退 |
| 推理 | batch/shape/CPU/CUDA | items/s、P95、内存 |
| 信号生成 | 单/多载波、批量 | samples/s、确定性 |
| 数据集 | HDF5/WebDataset、shuffle | samples/s、P95、缓存 |

## 2. 方法

Google Benchmark 固定构建和 CPU 电源策略；每项预热、至少 30 个样本或达到置信稳定，报告 P50/P95/最大值与置信区间。GPU 同步测端到端和 kernel，禁止只报最好值。结果含依赖锁、数据哈希、设备/驱动、后台负载和实际后端。

## 3. 产品指标

UI P95 ≤50 ms、交互 ≥30 FPS、热缓存三图一致结果 P95 ≤150 ms、冷缓存 ≤1 s、取消确认 P95 ≤200 ms、标准浏览 ≤4 GiB、隐藏视图 500 ms 内停止专属活动。所有指标按 B1 和 D1-D4 复核。""",
       "性能实现与验收", "SRS PERF、模块设计", "定义公共平台和应用端到端基准", common_refs)

    md("06_测试与验收/算法正确性验证方案.md", "SS-TEST-NUM-001", "算法正确性验证方案",
       """## 1. 参考原则

不自写 FFT/矩阵/滤波作为参考。黄金值来自经批准的独立成熟实现或解析信号；参考工具版本、参数、输入 SHA256 和生成脚本必须锁定。oneMKL、cuFFT、可选 FFTW 适配器相互交叉验证，但 GPL 参考制品不进入默认产品。

## 2. 数据与阈值

单音/双音/扫频/脉冲/噪声/直流/NaN/Inf/削顶；实/复、字节序、IQ/QI、块边界。时间 ≤1 样本；频率 ≤0.5 bin；PSD 幅度建议 ≤0.1 dB（待算法评审冻结）；CPU/GPU 使用 dtype 分级绝对/相对误差。

## 3. 算法

FFT/IFFT 往返与 Parseval；PSD 窗增益/ENBW/单位；STFT 帧时间、overlap、边界；滤波频响/时延/状态；重采样速率/混叠；频移相位；星座/眼图数据；模型前后处理和 Top-K。测试分块与整块等价、取消不发布、过期请求不提交。

## 4. 证据

每个结果保存 backend、version、device、input digest、parameter digest、output digest、tolerance、max/RMS error 和 pass/fail。图只作辅助，数值比较为判定依据。""",
       "算法实现与验收", "SRS NUM/DSP/VIS、依赖策略", "建立成熟参考与多后端一致性", common_refs,
       ["PSD 0.1 dB 阈值需信号算法、产品和测试联合批准。"])

    md("06_测试与验收/验收标准.md", "SS-TEST-ACCEPT-001", "验收标准",
       """## 1. 文档/平台基线验收

目录、元数据、编号、相对链接、架构 DAG、API 目录、Excel、SVG/PNG、许可证、依赖锁和追踪校验通过；每个文件均有实质内容。十模块职责、依赖、API、测试、发布和复用目标明确。

## 2. 软件 Release 门禁

1. P0 需求实测 100% 通过，P1 偏差有批准记录；
2. D1-D4/B1 的数值、性能、恢复、GPU 回退、插件故障和 8 小时耐久通过；
3. 公共包独立消费、Python wheel、SDK 示例、API/ABI、插件兼容和多应用并存通过；
4. Signal Generator 薄壳复用验证不复制 Data/DSP/Task/Visualization 实现；
5. 许可证/SBOM/源代码义务/安装回滚齐全，FFTW 不进入非 GPL 默认包；
6. 阻断、数据错误、来源错配、未授权网络连接为 0。

## 3. UI 门禁

1600×900/1366×768/1280×720 和 100–200% DPI 无关键截断；键盘可达、焦点可见、28px 命中；图表有范围/单位/质量文本摘要。P02 三图同窗、频率联动、隐藏调度、右键方向语义、色阶非计算语义通过。

## 4. 签署

产品、架构、信号算法、开发、测试、安全/许可证和交付共同签署。建议基线值必须先转为批准值或记录偏差。""",
       "验收与发布", "全部需求、测试、风险", "定义文档与软件两级门禁", common_refs)

    # Project plan.
    md("07_项目计划/开发里程碑.md", "SS-PM-MS-001", "开发里程碑",
       """## 里程碑

| ID | 目标 | 主要输出 | 退出条件 |
|---|---|---|---|
| MS-00 | 文档/架构/依赖基线 | ADR、API、锁、许可证、数据计划 | 架构与许可评审通过 |
| MS-01 | Core/Data/TaskRuntime | 公共包、RAW读取、任务/恢复 | 独立单测与包消费通过 |
| MS-02 | DSP/Compute | oneMKL CPU、CUDA适配骨架、黄金基准 | 数值与后端选择通过 |
| MS-03 | Visualization/Workbench | 公共图表、视口、Dock/中心/命令 | 1280×720/DPI/复用 demo |
| MS-04 | Signal Studio 基础 | 项目、导入、P01/P02、结果 | D0/D4流程、来源隔离 |
| MS-05 | 宽窄带与联动 | Selection、通道、DSP、Inspector | AT-05/14/23 |
| MS-06 | Plugin/Model/Dataset | SDK、ORT、数据集清单、示例插件 | 多宿主契约测试 |
| MS-07 | 工程化发布 | 性能、稳定、GPU矩阵、安装/SBOM | Beta 门禁 |
| MS-08 | Beta 多应用验证 | Signal Generator 薄壳、并存安装 | 无公共代码复制 |
| MS-09 | Release | Signal Studio 1.0 + SDK 1.0 | 全部 Release 门禁 |

## 演进

Phase 2 Signal Generator；Phase 3 Dataset Builder/Manager；Phase 4 Model Trainer/Evaluator；Phase 5 Inference Studio。各阶段只新增应用编排或领域服务，优先消费已有公共包；发现公共缺口需以 ADR/API 变更处理，不在应用私建副本。""",
       "项目实施", "平台总任务、WBS", "平台优先并加入第二宿主门禁", common_refs)

    md("07_项目计划/版本发布计划.md", "SS-PM-REL-001", "版本发布计划",
       """## 1. 发布列车

平台 SDK、插件 ABI、schema、应用分别版本化但使用兼容矩阵发布。每次列车包含 alpha（开发包）、beta（冻结 API）、rc（仅阻断修复）、release（签名包）和维护补丁。

## 2. 包与渠道

`signal-platform-runtime/dev/debug-symbols/python/licenses-sbom`、`signal-studio`、`signal-plugin-sdk`、示例插件。内部制品仓为权威来源；离线介质含哈希清单、依赖缓存、安装/回滚脚本。不同应用共享版本化 runtime，不共享可写配置。

## 3. 兼容矩阵

SDK 1.x ↔ Plugin ABI 1 ↔ Workspace 1.x；应用声明平台最小/最大次版本。破坏性 API/ABI/schema 变更升主版本并提供迁移与并行安装窗口。弃用至少一个次版本。

## 4. 发布门禁

可重复构建、符号/SBOM/许可证、病毒扫描、签名、包消费、wheel、插件、多应用并存、升级/回滚、无网络、设备矩阵和已知问题均通过。""",
       "配置与发布", "版本说明、ADR-020", "建立 SDK/应用/插件/schema 联合发布", common_refs)

    # Third-party and competitor evaluation.
    dep_rows = "\n".join(f"| {d['name']} | {d['version']} | {d['decision']} | {d['spdx']} | {d['lock']} | {d['notes']} |" for d in deps)
    md("08_参考资料/第三方库清单.md", "SS-DEP-001", "第三方库清单",
       f"""## 1. 决策原则

不自研成熟 FFT、矩阵、通用信号处理、日志、测试或模型运行时；功能重叠依赖不同时进入默认包；全部由 Adapter 隔离；公共 API 不暴露第三方类型。版本由 vcpkg baseline `82b6bc8…` 和外部 URL+SHA256 锁定。

## 2. 清单

| 库 | 锁定版本 | 决策 | 许可证/SPDX | 锁定机制 | 说明 |
|---|---|---|---|---|---|
{dep_rows}

## 3. 候选结论

- FFT：oneMKL 默认；cuFFT 可选；FFTW GPL 只用于 GPL/内部参考配置；
- 矩阵：Eigen 私有值计算，oneMKL 负责高性能内核；OpenBLAS 为替代候选；
- 图形：Qt Base + 可替换图表适配；Qt Graphs/QCustomPlot 必须先通过许可证和性能门禁；
- 通信 DSP：liquid-dsp 为 P1 适配器，不与基础滤波/FFT重复默认引入；
- 模型：ONNX Runtime 默认，LibTorch 延后，避免双运行时体积/冲突；
- 测试：GoogleTest + Google Benchmark；日志 spdlog + fmt。

## 4. 获取与离线

`依赖锁定/vcpkg.json`、`dependency-lock.json`、`scripts/fetch_dependencies.ps1` 提供固定来源和哈希。离线流程：联网机验证脚本 → `vcpkg install --binarysource=clear;files,<cache>,readwrite` → 保存 downloads/packages/binary cache/许可证/SBOM → 离线机只读校验后安装。CUDA 只下载已校验 network installer，安装由授权人员接受 EULA 后手动执行。""",
       "架构、构建、发布", "vcpkg 官方 baseline 与各项目官方来源", "锁定版本、许可证、获取与离线准备", common_refs,
       ["Qt 图表模块和 oneMKL 再分发条款需法律/采购在 MS-00 最终确认。"])

    md("08_参考资料/竞品参考/竞品能力对照.md", "SS-REF-COMP-001", "竞品能力对照",
       """## 1. 使用方式

本表仅作为需求发现框架，不复制界面、图标或受保护内容，也不宣称特定版本当前功能。采购或对外比较前必须在官方材料和实际试用中复核。

| 观察域 | 业界工具常见能力 | Signal Platform 取舍 |
|---|---|---|
| 频谱/时频 | 高动态范围、测量、标记、导出 | 公共 Visualization，整数 Hz 与来源链优先 |
| 大文件 | 分段/回放/索引 | LoadedDataRange、渐进瓦片、诚实质量状态 |
| 自动化 | 脚本/API/批处理 | Headless CLI + Python SDK + TaskRuntime |
| 算法扩展 | 插件/模型/解调 | 多宿主 Plugin SDK，UI 解耦、权限治理 |
| 数据与训练 | 数据集/标签/模型 | 独立 Dataset/Model 能力，不塞入 Studio |
| 部署 | 离线/许可证/硬件 | 锁定依赖、CPU 必选、GPU 可降级 |

## 2. 评审问题

所有对比必须回答：坐标与单位是否可信、结果是否可追溯、取消/失败是否留半成品、插件 ABI 是否稳定、公共能力能否被第二应用复用、离线和许可证义务是否清晰。""",
       "产品与架构调研", "公开行业经验；未绑定具体版本", "建立不依赖品牌素材的竞品评审框架", common_refs)


def write_delivery_catalog() -> None:
    files = sorted(p for p in ROOT.rglob("*") if p.is_file())
    rows = []
    for p in files:
        rel = p.relative_to(ROOT).as_posix()
        ext = p.suffix.lower().lstrip(".") or "file"
        layer = rel.split("/")[0]
        purpose = {
            "00_交付说明": "交付控制与评审", "01_需求": "需求与追踪", "02_原型设计": "交互与证据",
            "03_UI规范": "Design System 与品牌", "04_技术设计": "平台架构与 ADR", "05_接口与数据": "API/SDK/schema",
            "06_测试与验收": "测试资产", "07_项目计划": "里程碑/WBS/风险", "08_参考资料": "依赖与参考", "scripts": "生成、获取和校验",
        }.get(layer, "总览/索引")
        rows.append(f"| `{rel}` | {ext} | {p.stat().st_size} | {purpose} | 责任角色按所属域 | 基线评审/自动校验 |")
    body = f"""## 1. 阅读顺序

执行报告 → 需求与追踪 → 总体架构/ADR → API/SDK → 测试 → WBS/风险 → 依赖与资产。

## 2. 交付文件

| 相对路径 | 格式 | 字节 | 用途 | 责任角色 | 评审要求 |
|---|---|---:|---|---|---|
{chr(10).join(rows)}

## 3. 完整性规则

所有 Markdown 非空并含元数据/参考/未决/变更；Excel 可打开且含冻结、筛选、列宽、公式和说明；SVG 可解析并含许可证 metadata；PNG 可打开；HTML 保持可运行归档；所有引用使用相对路径。`document-index.json` 是机器可读索引，`validation-report.md` 是最终校验证据。

## 4. 层级归属

文档中的平台/基础库条目优先；Signal Studio 应用条目只组合工作流、页面、配置、权限、品牌和应用特定导出。"""
    md("00_交付说明/交付物清单.md", "SS-DEL-CAT-001", "交付物清单", body,
       "交付与评审", "本目录实际扫描", "生成逐文件用途和评审清单", ["本目录 `document-index.json`", "自动校验报告"])


def write_document_index() -> None:
    files = []
    for p in sorted(ROOT.rglob("*")):
        if not p.is_file():
            continue
        if p == ROOT / "document-index.json":
            continue
        data = p.read_bytes()
        files.append({"path": p.relative_to(ROOT).as_posix(), "bytes": len(data), "sha256": hashlib.sha256(data).hexdigest(), "extension": p.suffix.lower()})
    write_json(ROOT / "document-index.json", {"schema": "signal.delivery-index/1.0", "generated": DATE, "root": ".", "files": files})


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--logo-concept", type=Path, help="Optional AI concept raster to retain as design evidence")
    args = parser.parse_args()

    reqs = parse_requirements()
    pages = page_catalog()
    apis = api_catalog()
    tests = build_tests(reqs)
    libs = library_capabilities()
    copy_evidence()
    generate_assets(args.logo_concept)
    shutil.copy2(ROOT / "03_UI规范/Logo/SignalStudio_AppIcon.ico", ROOT / "02_原型设计/原型源文件/favicon.ico")
    generate_test_data_script_and_data()
    deps = dependency_files()
    requirements_workbook(reqs, tests)
    pages_workbook(pages)
    tests_workbook(tests)
    wbs_and_risk_workbooks(libs)
    generate_docs(reqs, pages, apis, tests, libs, deps)
    write_json(ROOT / "02_原型设计/page-catalog.json", {"schema": "signal.pages/1.0", "generated": DATE, "pages": pages})
    write_json(ROOT / "04_技术设计/library-catalog.json", {"schema": "signal.libraries/1.0", "generated": DATE, "capabilities": libs})
    write_json(ROOT / "06_测试与验收/test-catalog.json", {"schema": "signal.tests/1.0", "generated": DATE, "tests": tests})
    (ROOT / "02_原型设计/页面截图/标准截图").mkdir(parents=True, exist_ok=True)
    write_json(ROOT / "02_原型设计/页面截图/screenshot-manifest.json", {"schema": "signal.screenshots/1.0", "standard": [], "audit_count": len(list((ROOT / "02_原型设计/页面截图/audit").glob("*.png"))), "note": "标准截图由 Playwright 真实浏览器渲染后登记。"})
    write_delivery_catalog()
    write_document_index()
    print(json.dumps({"requirements": len(reqs), "tests": len(tests), "apis": len(apis), "libraries": 10, "pages": len(pages)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
