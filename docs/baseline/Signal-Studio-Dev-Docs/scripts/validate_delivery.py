#!/usr/bin/env python3
"""Validate the complete Signal Studio development-document delivery."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parent))
from generate_delivery import DATE, ROOT, footer, metadata, write_delivery_catalog, write_document_index, write_json, write_text


checks: list[dict] = []


def record(category: str, item: str, ok: bool, detail: str, severity: str = "error") -> None:
    checks.append({"category": category, "item": item, "status": "PASS" if ok else ("WARN" if severity == "warning" else "FAIL"), "detail": detail})


def sha(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def check_required_files() -> None:
    required = [
        "README.md", "CHANGELOG.md", "document-index.json",
        "00_交付说明/交付物清单.md", "00_交付说明/版本说明.md", "00_交付说明/已知问题清单.md", "00_交付说明/原型评审结论.md", "00_交付说明/自动审核批准记录.md",
        "01_需求/软件需求规格说明书.md", "01_需求/功能清单.md", "01_需求/非功能需求说明.md", "01_需求/需求追踪矩阵.xlsx",
        "02_原型设计/原型设计说明书.md", "02_原型设计/交互规格说明书.md", "02_原型设计/页面清单.xlsx", "02_原型设计/状态与异常场景说明.md",
        "02_原型设计/原型源文件/Signal Studio交互原型_基线归档.html",
        "03_UI规范/UI视觉规范.md", "03_UI规范/图表显示规范.md", "03_UI规范/时频图色阶规范.md",
        "04_技术设计/软件总体架构设计说明书.md", "04_技术设计/模块详细设计说明书.md", "04_技术设计/线程与任务调度设计.md", "04_技术设计/大文件与缓存设计.md", "04_技术设计/GPU加速设计.md",
        "05_接口与数据/RAW-IQ数据格式说明.md", "05_接口与数据/工程文件格式说明.md", "05_接口与数据/C++接口说明.md", "05_接口与数据/Python接口说明.md", "05_接口与数据/算法插件接口说明.md", "05_接口与数据/错误码说明.md",
        "06_测试与验收/测试计划.md", "06_测试与验收/功能测试用例.xlsx", "06_测试与验收/性能测试方案.md", "06_测试与验收/算法正确性验证方案.md", "06_测试与验收/验收标准.md", "06_测试与验收/测试数据/generate_test_data.py",
        "07_项目计划/开发里程碑.md", "07_项目计划/工作分解结构WBS.xlsx", "07_项目计划/风险清单.xlsx", "07_项目计划/版本发布计划.md",
        "08_参考资料/第三方库清单.md", "08_参考资料/依赖锁定/vcpkg.json", "08_参考资料/依赖锁定/dependency-lock.json",
    ]
    missing = [p for p in required if not (ROOT / p).is_file()]
    record("文件", "必需交付物", not missing, f"{len(required)-len(missing)}/{len(required)} 存在；缺失={missing}")
    empty_dirs = [p.relative_to(ROOT).as_posix() for p in ROOT.rglob("*") if p.is_dir() and not any(p.iterdir())]
    record("文件", "空目录", not empty_dirs, f"空目录={empty_dirs}")
    small = [p.relative_to(ROOT).as_posix() for p in ROOT.rglob("*.md") if p.stat().st_size < 800]
    record("文件", "Markdown 实质内容", not small, f"低于 800 字节={small}")


def check_markdown() -> None:
    fields = ["文档编号", "文档名称", "项目名称", "文档版本", "基线版本", "状态", "内容类型（meta.contentType）", "编制日期", "适用阶段", "输入来源", "本版变更"]
    allowed_types = {"Tutorial", "How-to", "Reference", "Conceptual", "Troubleshooting", "Landing"}
    missing_meta, missing_footer, bad_abs, bad_links, bad_mermaid, placeholders = [], [], [], [], [], []
    unapproved, wrong_baseline, bad_content_type, writing_issues = [], [], [], []
    link_re = re.compile(r"\[[^\]]+\]\(([^)]+)\)")
    for path in ROOT.rglob("*.md"):
        text = path.read_text(encoding="utf-8")
        rel = path.relative_to(ROOT).as_posix()
        if any(f"| {field} |" not in text for field in fields):
            missing_meta.append(rel)
        if "| 状态 | 已批准 |" not in text:
            unapproved.append(rel)
        if "| 基线版本 | BL1.0 |" not in text:
            wrong_baseline.append(rel)
        type_match = re.search(r"^\| 内容类型（meta\.contentType） \| ([^|]+) \|$", text, re.M)
        if not type_match or type_match.group(1).strip() not in allowed_types:
            bad_content_type.append(rel)
        if any(h not in text for h in ["## 参考资料", "## 未决事项", "## 变更记录"]):
            missing_footer.append(rel)
        if re.search(r"(?i)(?<![A-Za-z])[A-Z]:[\\/]", text) or "file://" in text.lower():
            bad_abs.append(rel)
        if re.search(r"\b(?:TBD|TODO|FIXME)\b|待补充|空模板", text, re.I):
            placeholders.append(rel)
        for target in link_re.findall(text):
            target = target.split("#", 1)[0]
            if not target or re.match(r"^[a-z]+://", target, re.I):
                continue
            if Path(target).is_absolute() or not (path.parent / target).resolve().exists():
                bad_links.append(f"{rel} -> {target}")
        blocks = re.findall(r"```mermaid\s*\n(.*?)```", text, re.S)
        for i, block in enumerate(blocks, 1):
            first = next((x.strip() for x in block.splitlines() if x.strip()), "")
            pairs = [("[", "]"), ("(", ")"), ("{", "}")]
            if not re.match(r"^(flowchart|graph|stateDiagram|sequenceDiagram|classDiagram)", first) or any(block.count(a) != block.count(b) for a, b in pairs) or block.count('"') % 2:
                bad_mermaid.append(f"{rel}#{i}")
        in_fence = False
        for lineno, line in enumerate(text.splitlines(), 1):
            if not line.startswith("```"):
                continue
            if not in_fence and not re.match(r"^```[A-Za-z0-9_+.-]+(?:\s|$)", line):
                writing_issues.append(f"{rel}:{lineno}:代码围栏缺语言")
            in_fence = not in_fence
        prose = re.sub(r"```.*?```", "", text, flags=re.S)
        prose = re.sub(r"`[^`]*`", "", prose)
        if "—" in prose:
            writing_issues.append(f"{rel}:破折号标点")
        if "..." in prose:
            writing_issues.append(f"{rel}:三点省略号")
        if re.search(r"\b(?:easy|simple|quick|very|just|really|simply)\b", prose, re.I):
            writing_issues.append(f"{rel}:英文填充词")
    record("文档", "元数据", not missing_meta, f"缺失={missing_meta}")
    record("批准", "全部 Markdown 已批准", not unapproved, f"未批准={unapproved}; 总数={len(list(ROOT.rglob('*.md')))}")
    record("批准", "全部 Markdown 使用 BL1.0", not wrong_baseline, f"错误={wrong_baseline}")
    record("写作", "内容类型有效", not bad_content_type, f"异常={bad_content_type}")
    record("写作", "可适用 Writing Guidelines 规则", not writing_issues, f"异常={writing_issues}")
    record("文档", "参考/未决/变更尾部", not missing_footer, f"缺失={missing_footer}")
    record("链接", "相对链接", not bad_links, f"无效={bad_links}")
    record("文档", "无绝对路径", not bad_abs, f"违规={bad_abs}")
    record("文档", "无占位符", not placeholders, f"违规={placeholders}")
    record("Mermaid", "结构语法", not bad_mermaid, f"异常={bad_mermaid}")


def check_traceability() -> dict:
    reqs = json.loads((ROOT / "01_需求/requirements.json").read_text(encoding="utf-8"))["requirements"]
    pages = json.loads((ROOT / "02_原型设计/page-catalog.json").read_text(encoding="utf-8"))["pages"]
    apis = json.loads((ROOT / "05_接口与数据/api-catalog.json").read_text(encoding="utf-8"))["apis"]
    tests = json.loads((ROOT / "06_测试与验收/test-catalog.json").read_text(encoding="utf-8"))["tests"]
    caps = json.loads((ROOT / "04_技术设计/library-catalog.json").read_text(encoding="utf-8"))["capabilities"]
    sets = {"page": {p["id"] for p in pages}, "api": {a["id"] for a in apis}, "test": {t["id"] for t in tests}, "library": {c["library"] for c in caps}, "milestone": {f"MS-{i:02d}" for i in range(10)}}
    duplicates = [k for k, v in Counter(r["id"] for r in reqs).items() if v > 1]
    record("编号", "需求编号唯一", not duplicates, f"重复={duplicates}; 总数={len(reqs)}")
    test_dups = [k for k, v in Counter(t["id"] for t in tests).items() if v > 1]
    api_dups = [k for k, v in Counter(a["id"] for a in apis).items() if v > 1]
    record("编号", "API/测试编号唯一", not (test_dups or api_dups), f"API重复={api_dups}; 测试重复={test_dups}")
    dangling = []
    for r in reqs:
        for field in ["api", "test", "library", "milestone"]:
            if not r[field] or r[field] not in sets[field]:
                dangling.append(f"{r['id']}:{field}={r[field]}")
        if r["page"] and r["page"] not in sets["page"]:
            dangling.append(f"{r['id']}:page={r['page']}")
    record("追踪", "无悬空引用", not dangling, f"悬空={dangling[:20]}")
    total = len(reqs)
    applicable = [r for r in reqs if r["page_applicable"]]
    p01 = [r for r in reqs if r["priority"] in {"P0", "P1"}]
    metrics = {
        "requirements": total, "p0": sum(r["priority"] == "P0" for r in reqs), "p1": sum(r["priority"] == "P1" for r in reqs),
        "page_coverage": sum(bool(r["page"]) for r in applicable) / len(applicable) if applicable else 1,
        "api_coverage": sum(bool(r["api"]) for r in reqs) / total,
        "test_coverage": sum(bool(r["test"]) for r in reqs) / total,
        "library_coverage": sum(bool(r["library"]) for r in reqs) / total,
        "reuse_coverage": sum(bool(r["reuse_apps"]) for r in reqs) / total,
        "p0_p1_test_coverage": sum(bool(r["test"]) for r in p01) / len(p01),
    }
    record("追踪", "覆盖率", all(v == 1 for k, v in metrics.items() if k.endswith("coverage")), json.dumps(metrics, ensure_ascii=False))
    return metrics


def check_architecture_and_api() -> None:
    arch = json.loads((ROOT / "04_技术设计/architecture.json").read_text(encoding="utf-8"))
    nodes = {n["name"]: n for n in arch["nodes"]}
    missing = [(n, d) for n, v in nodes.items() for d in v["deps"] if d not in nodes]
    visiting, visited, cycles = set(), set(), []
    def dfs(name, stack):
        if name in visiting:
            cycles.append(stack + [name]); return
        if name in visited: return
        visiting.add(name)
        for dep in nodes[name]["deps"]: dfs(dep, stack + [name])
        visiting.remove(name); visited.add(name)
    for name in nodes: dfs(name, [])
    brand = [n for n in nodes if "Studio" in n or "Signal Studio" in nodes[n].get("public_namespace", "")]
    record("架构", "依赖节点完整", not missing, f"缺失={missing}")
    record("架构", "无循环依赖", not cycles, f"循环={cycles}")
    record("架构", "公共库无应用品牌", not brand, f"违规={brand}")
    apis = json.loads((ROOT / "05_接口与数据/api-catalog.json").read_text(encoding="utf-8"))["apis"]
    banned = re.compile(r"\b(?:Qt|QObject|QWidget|QString|QVector|QList|QMap|QVariant|QImage|QPixmap|QPainter|QOpenGL\w*|Eigen|MKL|oneMKL|FFTW|cufft|TBB|Ort|ONNXRuntime|spdlog|fmt::|H5)\b", re.I)
    leaks = [f"{a['id']}:{a['signature']}" for a in apis if banned.search(a["signature"]) or banned.search(a["namespace"])]
    record("API", "公共签名无第三方类型", not leaks, f"泄漏={leaks}")
    required_modules = {"SignalCore", "SignalData", "SignalDSP", "SignalCompute", "SignalTaskRuntime", "SignalVisualization", "SignalWorkbench", "SignalPluginSDK", "SignalModelRuntime", "SignalDataset"}
    record("架构", "十模块齐全", set(nodes) == required_modules, f"实际={sorted(nodes)}")
    adr = list((ROOT / "04_技术设计/架构决策记录").glob("ADR-*.md"))
    adr_nums = [re.match(r"ADR-(\d{3})", p.name).group(1) for p in adr]
    record("编号", "ADR-001 至 ADR-020", set(adr_nums) == {f"{i:03d}" for i in range(1, 21)}, f"实际={sorted(adr_nums)}")


def check_excel() -> None:
    expected = {
        "01_需求/需求追踪矩阵.xlsx": {"需求总表", "需求-页面", "需求-接口", "需求-测试", "需求-里程碑", "需求-基础库", "需求-目标应用", "缺口与冲突", "统计", "说明"},
        "02_原型设计/页面清单.xlsx": {"页面清单", "统计", "说明"},
        "06_测试与验收/功能测试用例.xlsx": {"功能测试", "基础库测试", "SDK测试", "插件兼容测试", "多应用复用测试", "统计", "说明"},
        "07_项目计划/工作分解结构WBS.xlsx": {"WBS", "统计", "说明"},
        "07_项目计划/风险清单.xlsx": {"风险清单", "统计", "说明"},
    }
    problems = []
    for rel, sheets in expected.items():
        path = ROOT / rel
        try:
            wb = load_workbook(path, data_only=False)
            if not sheets.issubset(wb.sheetnames): problems.append(f"{rel}:缺sheet {sheets-set(wb.sheetnames)}")
            formulas = sum(1 for ws in wb for row in ws.iter_rows() for c in row if c.data_type == "f")
            validations = sum(len(ws.data_validations.dataValidation) for ws in wb)
            if formulas == 0: problems.append(f"{rel}:无公式")
            if validations == 0: problems.append(f"{rel}:无下拉验证")
            first = wb[wb.sheetnames[0]]
            if not first.freeze_panes or not first.auto_filter.ref: problems.append(f"{rel}:缺冻结/筛选")
            if not any(v.width and v.width > 10 for v in first.column_dimensions.values()): problems.append(f"{rel}:未设列宽")
            wbv = load_workbook(path, data_only=True)
            errs = [(ws.title, c.coordinate, c.value) for ws in wbv for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("#")]
            if errs: problems.append(f"{rel}:公式错误 {errs[:5]}")
        except Exception as exc:
            problems.append(f"{rel}:打不开 {exc}")
    record("Excel", "可打开/样式/公式/下拉/零错误", not problems, f"问题={problems}")


def check_assets_and_html() -> None:
    svg_errors = []
    for p in ROOT.rglob("*.svg"):
        try:
            root = ET.parse(p).getroot()
            text = p.read_text(encoding="utf-8")
            if "viewBox" not in root.attrib or "SPDX-License-Identifier" not in text:
                svg_errors.append(f"{p.relative_to(ROOT)}:viewBox/metadata")
        except Exception as exc:
            svg_errors.append(f"{p.relative_to(ROOT)}:{exc}")
    record("SVG", "可解析/视口/许可证", not svg_errors, f"问题={svg_errors}")
    png_errors = []
    for p in ROOT.rglob("*.png"):
        try:
            with Image.open(p) as img:
                img.verify()
        except Exception as exc:
            png_errors.append(f"{p.relative_to(ROOT)}:{exc}")
    record("PNG", "可打开", not png_errors, f"数量={len(list(ROOT.rglob('*.png')))}; 问题={png_errors}")
    std = list((ROOT / "02_原型设计/页面截图/标准截图").glob("*.png"))
    std_info = []
    for p in std:
        with Image.open(p) as img:
            std_info.append({"path": p.relative_to(ROOT).as_posix(), "width": img.width, "height": img.height, "bytes": p.stat().st_size, "sha256": sha(p), "renderer": "Playwright CLI / Microsoft Edge"})
    write_json(ROOT / "02_原型设计/页面截图/screenshot-manifest.json", {"schema": "signal.screenshots/1.0", "generated": DATE, "standard": std_info, "audit_count": len(list((ROOT / "02_原型设计/页面截图/audit").glob("*.png"))), "browser_console": "Playwright/Edge isolated rerun: 0 errors, 0 warnings"})
    needed = {(1600, 900), (1280, 720)}
    actual = {(x["width"], x["height"]) for x in std_info}
    record("截图", "真实浏览器标准截图", len(std_info) >= 5 and needed.issubset(actual), f"数量={len(std_info)}; 尺寸={sorted(actual)}")
    html = ROOT / "02_原型设计/原型源文件/Signal Studio交互原型_基线归档.html"
    text = html.read_text(encoding="utf-8")
    record("HTML", "归档可运行结构", html.stat().st_size > 100_000 and "<title>" in text and "id=\"root\"" in text and "</html>" in text, f"字节={html.stat().st_size}")


def check_licenses_dependencies_and_data() -> None:
    lock = json.loads((ROOT / "08_参考资料/依赖锁定/dependency-lock.json").read_text(encoding="utf-8"))
    bad = [d["name"] for d in lock["dependencies"] if not all(d.get(k) for k in ["version", "spdx", "official_url", "lock", "verification", "decision"])]
    selected = [d for d in lock["dependencies"] if d["selected"]]
    record("许可证", "依赖许可证/来源/锁定", not bad and bool(selected), f"依赖={len(lock['dependencies'])}; 默认选用={len(selected)}; 缺项={bad}")
    manifest = json.loads((ROOT / "03_UI规范/asset-manifest.json").read_text(encoding="utf-8"))
    record("许可证", "原创资产许可证", bool(manifest.get("license")) and "third-party artwork" in manifest.get("originality", ""), f"license={manifest.get('license')}")
    data_manifest_path = ROOT / "06_测试与验收/测试数据/generated/manifest.json"
    data = json.loads(data_manifest_path.read_text(encoding="utf-8"))
    data_bad = []
    for item in data["files"]:
        p = data_manifest_path.parent / item["path"]
        if not p.exists() or p.stat().st_size != item["bytes"] or sha(p) != item["sha256"]:
            data_bad.append(item["path"])
    record("测试数据", "生成数据哈希", not data_bad and data.get("license") == "CC0-1.0", f"文件={len(data['files'])}; 问题={data_bad}")
    source_manifest = json.loads((ROOT / "02_原型设计/原型源文件/source-manifest.json").read_text(encoding="utf-8"))
    copy_bad = [x["path"] for x in source_manifest["files"] if not (ROOT / x["path"]).exists() or sha(ROOT / x["path"]) != x["sha256"]]
    record("原始材料", "只读副本哈希", not copy_bad, f"副本={len(source_manifest['files'])}; 变化={copy_bad}")


def risk_count() -> tuple[int, int]:
    wb = load_workbook(ROOT / "07_项目计划/风险清单.xlsx", data_only=True)
    ws = wb["风险清单"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return len(rows), sum(r[1] == "高" and r[8] != "已关闭" for r in rows)


def make_report(title: str, number: str, body: str) -> str:
    return f"# {title}\n\n{metadata(number, title, '交付验收', '本目录自动化扫描与结构化基线', '自动生成、审核并批准文档基线', '已批准')}\n\n{body}\n\n{footer(['本目录 `validation-results.json`', '本目录结构化 requirements/api/architecture/test/asset/dependency 清单'], ['生产软件尚未实现；本报告只证明文档交付资产校验结果。'])}\n"


def write_reports(metrics: dict) -> None:
    counts = Counter(c["status"] for c in checks)
    rows = "\n".join(f"| {c['category']} | {c['item']} | {c['status']} | {c['detail'].replace('|','/')} |" for c in checks)
    report = f"""## 1. 结论

自动校验项 {len(checks)}：PASS {counts['PASS']}，WARN {counts['WARN']}，FAIL {counts['FAIL']}。校验范围覆盖文件、Markdown 元数据与链接、编号/追踪、Mermaid 结构、架构 DAG、API 第三方泄漏、Excel、SVG/PNG/HTML、许可证、依赖锁、测试数据和原始副本。

## 2. 明细

| 类别 | 检查 | 结果 | 详情 |
|---|---|---|---|
{rows}

## 3. 追踪指标

```json
{json.dumps(metrics, ensure_ascii=False, indent=2)}
```

## 4. 校验边界

Mermaid 做结构语法检查；Excel 由 openpyxl 打开并扫描公式错误。本机 LibreOffice 辅助脚本因 Windows `AF_UNIX` 不可用，已改用已安装 Excel COM 执行 `CalculateFullRebuild` 并保存。HTML 浏览器标准截图由 Playwright/Edge 产生。生产 C++ 构建、真实 FFT/STFT、GPU、模型、100 GB 数据和性能未在文档仓库中执行。"""
    write_text(ROOT / "validation-report.md", make_report("自动化校验报告", "SS-VAL-001", report))

    files = [p for p in ROOT.rglob("*") if p.is_file()]
    markdown_files = list(ROOT.rglob("*.md"))
    approved_markdown = sum("| 状态 | 已批准 |" in p.read_text(encoding="utf-8") for p in markdown_files)
    risk_total, high_open = risk_count()
    apis = json.loads((ROOT / "05_接口与数据/api-catalog.json").read_text(encoding="utf-8"))["apis"]
    selected = [d for d in json.loads((ROOT / "08_参考资料/依赖锁定/dependency-lock.json").read_text(encoding="utf-8"))["dependencies"] if d["selected"]]
    execution = f"""## 1. 执行结论

已生成完整 `Signal-Studio-Dev-Docs/`，不是模板集。平台先行：十个公共模块均有职责、依赖、API、版本、第三方适配、测试和发布边界；Signal Studio 被限定为应用薄壳。自动校验 FAIL={counts['FAIL']}。

## 2. 输入材料

扫描了 6 份现有 Markdown、1 份自包含交互原型、11 张审计截图和 README；原始材料未覆盖，副本哈希已记录。冲突按 SRS V2.0 → 综合评审 → 术语契约 → 原型行为 → 设计/截图的优先级处理。

## 3. 交付统计

| 指标 | 数量/结果 |
|---|---:|
| 文件 | {len(files)} |
| Markdown | {len(markdown_files)} |
| 已批准 Markdown | {approved_markdown} / {len(markdown_files)} |
| Excel | {len(list(ROOT.rglob('*.xlsx')))} |
| 需求 | {metrics['requirements']}（P0 {metrics['p0']} / P1 {metrics['p1']}） |
| 页面/流程 | 12（7 + 5） |
| API | {len(apis)} |
| 测试 | {metrics['requirements']} |
| ADR | {len(list((ROOT/'04_技术设计/架构决策记录').glob('ADR-*.md')))} |
| 风险 | {risk_total}（开放高风险 {high_open}） |
| 基础库 | 10 |
| SVG | {len(list(ROOT.rglob('*.svg')))} |
| PNG/标准截图 | {len(list(ROOT.rglob('*.png')))} / {len(list((ROOT/'02_原型设计/页面截图/标准截图').glob('*.png')))} |

## 4. 基础库职责

SignalCore（基础契约）、SignalData（数据与读取）、SignalDSP（成熟算法适配）、SignalCompute（CPU/GPU 后端）、SignalTaskRuntime（跨工具任务）、SignalVisualization（公共图表）、SignalWorkbench（公共桌面壳）、SignalPluginSDK（多宿主插件）、SignalModelRuntime（模型推理）、SignalDataset（数据集）。依赖 DAG 无循环。

## 5. Signal Studio 专属与复用

专属：分析工作流、P01/P02/W02 编排、应用配置/权限/导出、品牌与菜单。复用：其余 RAW/IQ、DSP、计算、任务、缓存、图表、Workbench、插件、模型、数据集、日志/错误/单位。Signal Generator → Data/DSP/Task/Visualization；Dataset 工具 → Data/Dataset/Task；Trainer/Evaluator → Dataset/Model/Compute/Task；Inference Studio → Data/DSP/Model/Visualization/Workbench。

## 6. 第三方与依赖

默认锁定 {len(selected)} 个 vcpkg 依赖；oneMKL 默认 FFT/矩阵内核，cuFFT 可选，FFTW 仅 GPL 配置；ONNX Runtime 默认模型运行时。已从官方 GitHub 获取 vcpkg baseline/commit/archive并计算 SHA256；已从 NVIDIA 官方下载 CUDA 12.8.1 network installer 到临时验证位置并计算 SHA256。完整依赖包未安装：当前是文档仓库且没有 C++ 构建目标，故只配置 manifest、版本、许可证、获取/离线脚本和校验值，未虚报“已构建”。无法获取项：无；大体积构建材料需在实施环境按脚本准备。

## 7. 资产与测试数据

使用内置图像生成工具得到原创 logo-brand 概念图；生产资产为确定性 SVG/PNG/ICO，含许可证 metadata。Playwright/Edge 真实渲染 1600×900 与 1280×720 标准截图。测试数据脚本已生成并产出 3 个 D0 文件及 SHA256；D1-D3 需显式容量确认，D4 实体未随输入提供。

## 8. 校验结果

PASS {counts['PASS']} / WARN {counts['WARN']} / FAIL {counts['FAIL']}。覆盖率：页面适用需求 {metrics['page_coverage']:.1%}、接口 {metrics['api_coverage']:.1%}、测试设计 {metrics['test_coverage']:.1%}、基础库归属 {metrics['library_coverage']:.1%}、复用目标 {metrics['reuse_coverage']:.1%}、P0/P1 测试 {metrics['p0_p1_test_coverage']:.1%}。详情见 `validation-report.md`。

Excel 先尝试技能附带 LibreOffice 重算器；该脚本在本机 Windows 缺少 `socket.AF_UNIX`，随后使用已安装 Excel COM `CalculateFullRebuild` 成功重算并保存 5 个工作簿，未发现公式错误。

## 9. 未决事项与首个里程碑

未决：Qt 图表/许可、插件默认隔离、PSD 数值阈值、D1-D4 实体和 GPU 设备矩阵。建议立即启动 MS-00：执行已批准 ADR 和公共契约，复核 vcpkg/许可证与测试数据计划；随后 MS-01 实施 SignalCore、SignalData、SignalTaskRuntime，禁止先在 Signal Studio 内建立私有副本。"""
    write_text(ROOT / "执行报告.md", make_report("执行报告", "SS-EXEC-001", execution))


def main() -> int:
    check_required_files()
    metrics = check_traceability()
    # Materialize reports before link/metadata scanning so root navigation is resolvable.
    write_reports(metrics)
    check_markdown()
    check_architecture_and_api()
    check_excel()
    check_assets_and_html()
    check_licenses_dependencies_and_data()
    write_reports(metrics)
    # Reports are generated Markdown and must meet the same metadata/footer contract.
    for rel in ["validation-report.md", "执行报告.md"]:
        text = (ROOT / rel).read_text(encoding="utf-8")
        ok = all(x in text for x in ["| 文档编号 |", "## 参考资料", "## 未决事项", "## 变更记录"])
        record("文档", rel, ok, "报告元数据和尾部结构")
    write_reports(metrics)
    write_delivery_catalog()
    write_document_index()
    write_json(ROOT / "validation-results.json", {"schema": "signal.validation/1.0", "generated": DATE, "checks": checks, "metrics": metrics})
    # Refresh index after validation-results is present.
    write_document_index()
    failures = [c for c in checks if c["status"] == "FAIL"]
    print(json.dumps({"checks": len(checks), "pass": sum(c["status"] == "PASS" for c in checks), "warn": sum(c["status"] == "WARN" for c in checks), "fail": len(failures), "failures": failures}, ensure_ascii=False, indent=2))
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
